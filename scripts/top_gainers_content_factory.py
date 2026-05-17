#!/usr/bin/env python3
"""
Automated Top Gainers Content Factory (Dropbox & CSV Edition).

Reads the master spreadsheet config strictly from cell P2, fetches top gainers,
creates 1080x1350 finance images, uploads them to Dropbox, and writes
content_factory_post.csv for batch posting.
"""

from __future__ import annotations

import argparse
import csv
import functools
import hashlib
import io
import json
import math
import os
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlparse


ROOT = Path(__file__).resolve().parents[1]
FONT_DIR = ROOT / "public" / "Font_stock"
DEFAULT_DROPBOX_ROOT = "/Stock_Gainers_Content"
DEFAULT_OUTPUT_ROOT = ROOT / "temp"
DEFAULT_CANVAS_SIZE = (1080, 1350)
SQUARE_CANVAS_SIZE = (1080, 1080)
CANVAS_SIZE = DEFAULT_CANVAS_SIZE
BG = "#0d1117"
BULL = "#3fb950"
BEAR = "#f85149"
MUTED = "#8b949e"
TEXT = "#f0f6fc"
VIRAL_HEADLINE_FONT = "Mitr-Medium.ttf"


HEADLINE_COLOR_THEMES = {
    "classic": {
        "name": "Classic Red Blue",
        "line1_bg": "#e11d1d",
        "line1_text": "#ffffff",
        "line2_text": "#ffffff",
        "line2_highlight": "#fff200",
        "line3_bg": "#1688f0",
        "line3_text": "#ffffff",
    },
    "emerald_gold": {
        "name": "Emerald Gold",
        "line1_bg": "#059669",
        "line1_text": "#ffffff",
        "line2_text": "#ffffff",
        "line2_highlight": "#facc15",
        "line3_bg": "#0f766e",
        "line3_text": "#ffffff",
    },
    "orange_teal": {
        "name": "Orange Teal",
        "line1_bg": "#f97316",
        "line1_text": "#ffffff",
        "line2_text": "#ffffff",
        "line2_highlight": "#fde047",
        "line3_bg": "#0891b2",
        "line3_text": "#ffffff",
    },
    "purple_lime": {
        "name": "Purple Lime",
        "line1_bg": "#7c3aed",
        "line1_text": "#ffffff",
        "line2_text": "#ffffff",
        "line2_highlight": "#bef264",
        "line3_bg": "#4f46e5",
        "line3_text": "#ffffff",
    },
    "rose_cyan": {
        "name": "Rose Cyan",
        "line1_bg": "#e11d48",
        "line1_text": "#ffffff",
        "line2_text": "#ffffff",
        "line2_highlight": "#22d3ee",
        "line3_bg": "#0e7490",
        "line3_text": "#ffffff",
    },
    "amber_indigo": {
        "name": "Amber Indigo",
        "line1_bg": "#d97706",
        "line1_text": "#ffffff",
        "line2_text": "#ffffff",
        "line2_highlight": "#fef08a",
        "line3_bg": "#4f46e5",
        "line3_text": "#ffffff",
    },
    "magenta_mint": {
        "name": "Magenta Mint",
        "line1_bg": "#db2777",
        "line1_text": "#ffffff",
        "line2_text": "#ffffff",
        "line2_highlight": "#a7f3d0",
        "line3_bg": "#10b981",
        "line3_text": "#04111d",
    },
    "graphite_gold": {
        "name": "Graphite Gold",
        "line1_bg": "#374151",
        "line1_text": "#ffffff",
        "line2_text": "#ffffff",
        "line2_highlight": "#fbbf24",
        "line3_bg": "#92400e",
        "line3_text": "#ffffff",
    },
    "navy_coral": {
        "name": "Navy Coral",
        "line1_bg": "#1d4ed8",
        "line1_text": "#ffffff",
        "line2_text": "#ffffff",
        "line2_highlight": "#fb7185",
        "line3_bg": "#0f172a",
        "line3_text": "#ffffff",
    },
    "white_hot": {
        "name": "White Hot",
        "line1_bg": "#f8fafc",
        "line1_text": "#111827",
        "line2_text": "#ffffff",
        "line2_highlight": "#fb923c",
        "line3_bg": "#facc15",
        "line3_text": "#111827",
    },
}


@dataclass
class DropboxCreds:
    access_token: str = ""
    refresh_token: str = ""
    app_key: str = ""
    app_secret: str = ""


@dataclass
class AppCreds:
    openrouter_key: str = ""
    openai_key: str = ""
    news_api_key: str = ""
    giphy_api_key: str = ""
    dropbox: DropboxCreds = field(default_factory=DropboxCreds)


@dataclass
class StockItem:
    symbol: str
    company_name: str = ""
    price: float = 0.0
    pct_change: float = 0.0
    volume: int = 0
    sector: str = ""
    industry: str = ""
    website: str = ""
    domain: str = ""
    ohlc: Any = None
    news: list[str] = field(default_factory=list)
    headline: str = ""
    caption: str = ""
    image_path: Path | None = None
    dropbox_url: str = ""
    pe_ratio: float = 0.0
    news_items: list["NewsItem"] = field(default_factory=list)


@dataclass
class NewsItem:
    title: str = ""
    source: str = ""
    url: str = ""
    image_url: str = ""
    image_bytes: bytes | None = None


def load_dotenv_if_available() -> None:
    try:
        from dotenv import load_dotenv

        load_dotenv(ROOT / ".env")
        load_dotenv()
    except Exception:
        pass


def read_profiles() -> list[dict[str, Any]]:
    path = ROOT / "public" / "app_data" / "api_profiles.json"
    if not path.exists():
        return []
    try:
        data = json.loads(path.read_text("utf-8"))
        return data if isinstance(data, list) else []
    except Exception:
        return []


def load_credentials(profile_name: str = "") -> AppCreds:
    load_dotenv_if_available()
    profiles = read_profiles()
    profile = None
    if profile_name:
        profile = next((p for p in profiles if p.get("name") == profile_name or p.get("id") == profile_name), None)
    if profile is None:
        profile = next((p for p in profiles if p.get("active")), None) or (profiles[0] if profiles else {})

    return AppCreds(
        openrouter_key=os.getenv("OPENROUTER_API_KEY") or profile.get("openRouterKey", ""),
        openai_key=os.getenv("OPENAI_API_KEY", ""),
        news_api_key=os.getenv("NEWS_API_KEY", ""),
        giphy_api_key=os.getenv("GIPHY_API_KEY") or profile.get("giphyKey", ""),
        dropbox=DropboxCreds(
            access_token=os.getenv("DROPBOX_ACCESS_TOKEN") or profile.get("dropboxKey", ""),
            refresh_token=os.getenv("DROPBOX_REFRESH_TOKEN") or profile.get("dropboxRefreshToken", ""),
            app_key=os.getenv("DROPBOX_APP_KEY") or profile.get("dropboxAppKey", ""),
            app_secret=os.getenv("DROPBOX_APP_SECRET") or profile.get("dropboxAppSecret", ""),
        ),
    )


def read_p2(spreadsheet_path: Path, sheet_name: str = "") -> str:
    suffix = spreadsheet_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        from openpyxl import load_workbook

        wb = load_workbook(spreadsheet_path, data_only=True, read_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active
        value = ws["P2"].value
        return "" if value is None else str(value).strip()

    if suffix == ".csv":
        with spreadsheet_path.open("r", encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))
        if len(rows) < 2 or len(rows[1]) < 16:
            return ""
        return str(rows[1][15]).strip()

    raise ValueError(f"Unsupported spreadsheet type: {spreadsheet_path.suffix}. Use .xlsx, .xlsm, or .csv.")


def parse_p2_config(raw: str) -> dict[str, Any]:
    clean = raw.strip()
    lowered = clean.lower()
    symbols: list[str] = []
    sector = ""
    all_market_values = {
        "",
        "all",
        "all market",
        "all markets",
        "all sector",
        "all sectors",
        "market: all",
        "market=all",
        "sector: all",
        "sector=all",
        "sector: all sectors",
        "sector=all sectors",
        "ทั้งตลาด",
        "ทุกอุตสาหกรรม",
        "ทั้งตลาด / ทุกอุตสาหกรรม",
    }
    if lowered in all_market_values:
        return {"raw": raw, "symbols": [], "sector": ""}

    sector_match = re.search(r"(?:sector|target sector)\s*[:=]\s*([^;\n]+)", clean, re.I)
    if sector_match:
        sector = sector_match.group(1).strip()
        if sector.lower() in {"all", "all sector", "all sectors", "ทุกอุตสาหกรรม", "ทั้งตลาด"}:
            sector = ""

    symbol_match = re.search(r"(?:symbols?|tickers?|watchlist|override symbols?)\s*[:=]\s*([^;\n]+)", clean, re.I)
    if symbol_match:
        symbols = extract_symbol_tokens(symbol_match.group(1), allow_lowercase=True)
    elif not sector_match:
        symbols = extract_symbol_tokens(clean, allow_lowercase=False)

    if not sector and lowered.startswith("sector:"):
        sector = clean.split(":", 1)[1].strip()

    # If the whole P2 looks like prose rather than tickers, treat it as a sector/filter hint.
    if not symbols and not sector and clean:
        sector = clean

    return {"raw": raw, "symbols": list(dict.fromkeys(symbols)), "sector": sector}


def extract_symbol_tokens(value: str, allow_lowercase: bool) -> list[str]:
    symbols: list[str] = []
    for raw_token in re.split(r"[,;\s]+", value):
        raw_token = raw_token.strip()
        token = raw_token.upper()
        if token in {"SECTOR", "WATCHLIST", "SYMBOL", "SYMBOLS", "TICKER", "TICKERS"}:
            continue
        if not re.fullmatch(r"[A-Z][A-Z0-9.\-]{0,8}", token):
            continue
        if not allow_lowercase and raw_token != token:
            continue
        symbols.append(token)
    return symbols


def requests_session():
    import requests

    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": "BulkVideoCreatorApp/TopGainersContentFactory",
            "Accept": "application/json,text/html;q=0.8,*/*;q=0.5",
        }
    )
    return session


def fetch_yahoo_screener(scr_id: str, count: int) -> list[dict[str, Any]]:
    session = requests_session()
    url = "https://query1.finance.yahoo.com/v1/finance/screener/predefined/saved"
    # Yahoo Finance predefined screeners have a maximum limit of 250 items.
    params = {"scrIds": scr_id, "count": min(max(count, 25), 250)}
    res = session.get(url, params=params, timeout=20)
    res.raise_for_status()
    payload = res.json()
    quotes = payload.get("finance", {}).get("result", [{}])[0].get("quotes", [])
    return quotes if isinstance(quotes, list) else []


def fetch_yahoo_day_gainers(count: int) -> list[dict[str, Any]]:
    return fetch_yahoo_screener("day_gainers", count)


SECTOR_FALLBACK_SYMBOLS = {
    "technology": ["NVDA", "MSFT", "AAPL", "AVGO", "AMD", "ORCL", "CRM", "ADBE", "CSCO", "NOW", "INTC", "MU", "QCOM", "PLTR", "SNOW"],
    "healthcare": ["LLY", "UNH", "JNJ", "ABBV", "MRK", "TMO", "ABT", "ISRG", "AMGN", "PFE", "GILD", "REGN", "VRTX", "BMY", "ZTS"],
    "financial services": ["JPM", "BAC", "WFC", "GS", "MS", "C", "BLK", "AXP", "SCHW", "USB", "PNC", "COF", "BK", "TFC", "AFL"],
    "consumer cyclical": ["AMZN", "TSLA", "HD", "MCD", "NKE", "SBUX", "LOW", "BKNG", "TJX", "MAR", "ORLY", "ABNB", "RCL", "GM", "F"],
    "communication services": ["GOOGL", "META", "NFLX", "DIS", "TMUS", "VZ", "T", "CMCSA", "CHTR", "SPOT", "EA", "TTWO", "SNAP", "PINS", "ROKU"],
    "industrials": ["GE", "CAT", "RTX", "UNP", "HON", "UPS", "BA", "LMT", "DE", "ADP", "ETN", "PH", "WM", "EMR", "FDX"],
    "energy": ["XOM", "CVX", "COP", "SLB", "EOG", "MPC", "PSX", "VLO", "OXY", "HAL", "BKR", "DVN", "FANG", "APA", "HES"],
    "basic materials": ["LIN", "FCX", "NEM", "SCCO", "DD", "DOW", "APD", "ECL", "NUE", "STLD", "ALB", "MOS", "CF", "CLF", "CDE", "AA", "X"],
    "real estate": ["PLD", "AMT", "EQIX", "WELL", "SPG", "O", "DLR", "PSA", "CCI", "CBRE", "VICI", "EXR", "AVB", "EQR", "WY"],
    "utilities": ["NEE", "SO", "DUK", "CEG", "AEP", "SRE", "D", "EXC", "XEL", "PEG", "ED", "WEC", "PCG", "DTE", "EIX"],
}


def sector_fallback_symbols(sector: str) -> list[str]:
    lowered = sector.strip().lower()
    for key, symbols in SECTOR_FALLBACK_SYMBOLS.items():
        if key in lowered or lowered in key:
            return symbols
    return []


def fetch_symbol_items(symbols: list[str], *, sector_filter: str = "", strict_sector: bool = True) -> list[StockItem]:
    items: list[StockItem] = []
    seen: set[str] = set()
    for symbol in symbols:
        symbol = symbol.upper().strip()
        if not symbol or symbol in seen:
            continue
        seen.add(symbol)
        try:
            item = fetch_history_and_info(symbol)
            if strict_sector and sector_filter and sector_filter not in item.sector.lower():
                continue
            items.append(item)
        except Exception as exc:
            print(f"[WARN] Skip {symbol}: {exc}", file=sys.stderr)
    return items


def _quote_value(quote: dict[str, Any] | None, *keys: str, default: Any = None) -> Any:
    if not quote:
        return default
    for key in keys:
        value = quote.get(key)
        if value is not None and value != "":
            return value
    return default


def _quote_float(quote: dict[str, Any] | None, *keys: str, default: float = 0.0) -> float:
    value = _quote_value(quote, *keys)
    try:
        return float(value)
    except Exception:
        return default


def _quote_int(quote: dict[str, Any] | None, *keys: str, default: int = 0) -> int:
    value = _quote_value(quote, *keys)
    try:
        return int(float(value))
    except Exception:
        return default


def fetch_history_and_info(symbol: str, quote: dict[str, Any] | None = None, *, fetch_profile: bool = True) -> StockItem:
    import yfinance as yf

    ticker = yf.Ticker(symbol)
    hist = ticker.history(period="30d", interval="1d", auto_adjust=False)
    if hist is None or hist.empty:
        raise ValueError(f"No 30-day OHLC data returned for {symbol}")

    info: dict[str, Any] = {}
    if fetch_profile:
        try:
            info = ticker.get_info() or {}
        except Exception:
            try:
                info = ticker.info or {}
            except Exception:
                info = {}

    latest = hist.iloc[-1]
    prev_close = float(hist["Close"].iloc[-2]) if len(hist) >= 2 else float(latest["Open"])
    price = float(info.get("regularMarketPrice") or _quote_float(quote, "regularMarketPrice", "regularMarketPreviousClose", default=float(latest["Close"])))
    pct_change = info.get("regularMarketChangePercent")
    if pct_change is None:
        pct_change = _quote_value(quote, "regularMarketChangePercent", "regularMarketPercentChange")
    if pct_change is None and prev_close:
        pct_change = ((price - prev_close) / prev_close) * 100

    website = str(info.get("website") or "")
    domain = domain_from_url(website)
    pe_raw = info.get("trailingPE") or info.get("forwardPE")
    pe_ratio = float(pe_raw) if pe_raw is not None else 0.0
    quote_name = str(
        _quote_value(
            quote,
            "shortName",
            "longName",
            "displayName",
            "symbol",
            default=symbol.upper(),
        )
    )
    return StockItem(
        symbol=symbol.upper(),
        company_name=str(info.get("shortName") or info.get("longName") or quote_name or symbol.upper()),
        price=float(info.get("regularMarketPrice") or _quote_float(quote, "regularMarketPrice", "regularMarketPreviousClose", default=float(latest["Close"]))),
        pct_change=float(pct_change or 0),
        volume=int(info.get("regularMarketVolume") or _quote_int(quote, "regularMarketVolume", "averageDailyVolume3Month", default=int(latest.get("Volume") or 0))),
        sector=str(info.get("sector") or _quote_value(quote, "sector", default="") or ""),
        industry=str(info.get("industry") or _quote_value(quote, "industry", default="") or ""),
        website=website,
        domain=domain,
        ohlc=hist,
        pe_ratio=pe_ratio,
    )


def get_stock_universe(config: dict[str, Any], limit: int, scan_count: int) -> list[StockItem]:
    symbols = config["symbols"]
    if not symbols:
        quotes = fetch_yahoo_day_gainers(max(scan_count, limit * 20))
        sector_filter = str(config.get("sector") or "").strip().lower()
        symbols = [str(q.get("symbol", "")).upper() for q in quotes if q.get("symbol")]
        if not symbols:
            raise RuntimeError("Yahoo screener did not return any day gainers.")

        print(f"[INFO] Yahoo day gainers scanned: {len(symbols)} candidates")
        if sector_filter:
            print(f"[INFO] Sector filter from P2: {config.get('sector')}")

        items: list[StockItem] = []
        for symbol in symbols:
            try:
                item = fetch_history_and_info(symbol)
                if sector_filter and sector_filter not in item.sector.lower():
                    continue
                items.append(item)
                if len(items) >= limit:
                    break
            except Exception as exc:
                print(f"[WARN] Skip {symbol}: {exc}", file=sys.stderr)
        if sector_filter and len(items) < limit:
            print(
                f"[WARN] Found only {len(items)} matching {config.get('sector')} gainers "
                f"after scanning {len(symbols)} Yahoo candidates.",
                file=sys.stderr,
            )
            fallback_symbols = [s for s in sector_fallback_symbols(str(config.get("sector") or "")) if s not in symbols]
            if fallback_symbols:
                print(
                    f"[INFO] Sector fallback watchlist: scanning {len(fallback_symbols)} {config.get('sector')} symbols"
                )
                fallback_items = fetch_symbol_items(fallback_symbols, sector_filter=sector_filter, strict_sector=False)
                positive_items = [item for item in fallback_items if item.pct_change > 0]
                ranked_fallback = sorted(positive_items or fallback_items, key=lambda i: i.pct_change, reverse=True)
                if ranked_fallback:
                    if not positive_items:
                        print(
                            f"[WARN] No positive {config.get('sector')} movers found in fallback; using strongest available sector stocks.",
                            file=sys.stderr,
                        )
                    items.extend(ranked_fallback)
                    items = sorted(
                        list({item.symbol: item for item in items}.values()),
                        key=lambda i: i.pct_change,
                        reverse=True,
                    )[:limit]
        return sorted(items, key=lambda i: i.pct_change, reverse=True)[:limit]

    items = []
    for symbol in symbols[:limit]:
        try:
            items.append(fetch_history_and_info(symbol))
        except Exception as exc:
            print(f"[WARN] Skip {symbol}: {exc}", file=sys.stderr)
    return sorted(items, key=lambda i: i.pct_change, reverse=True)[:limit]


def _screen_stocks_from_quotes(quotes: list[dict[str, Any]], limit: int,
                               sector_filter: str = "", sort_key=None, reverse: bool = True,
                               fetch_profile: bool = True) -> list[StockItem]:
    quote_by_symbol = {str(q.get("symbol", "")).upper(): q for q in quotes if q.get("symbol")}
    symbols = list(quote_by_symbol.keys())
    items: list[StockItem] = []
    for scan_index, symbol in enumerate(symbols, start=1):
        try:
            print(f"[INFO] Screening {symbol} ({scan_index}/{len(symbols)})")
            item = fetch_history_and_info(symbol, quote_by_symbol.get(symbol), fetch_profile=fetch_profile)
            if sector_filter and sector_filter not in item.sector.lower():
                continue
            items.append(item)
            print(f"[INFO] Selected {item.symbol}: {item.pct_change:+.2f}% volume {item.volume:,} ({len(items)}/{limit})")
            if len(items) >= limit:
                break
        except Exception as exc:
            print(f"[WARN] Skip {symbol}: {exc}", file=sys.stderr)
    if sort_key:
        items = sorted(items, key=sort_key, reverse=reverse)
    return items[:limit]


def _screen_low_pe(limit: int, scan_count: int, pe_threshold: float = 20.0) -> list[StockItem]:
    # Try the undervalued screener first; fall back to most_actives
    quotes = fetch_yahoo_screener("undervalued_large_caps", max(scan_count, limit * 20))
    if not quotes:
        quotes = fetch_yahoo_screener("most_actives", max(scan_count, limit * 20))
    if not quotes:
        raise RuntimeError("Yahoo screener returned no candidates for low P/E screening.")

    symbols = [str(q.get("symbol", "")).upper() for q in quotes if q.get("symbol")]
    print(f"[INFO] Low P/E screening: {len(symbols)} candidates, PE threshold < {pe_threshold}")
    items: list[StockItem] = []
    all_pe_items: list[StockItem] = []
    for symbol in symbols:
        try:
            item = fetch_history_and_info(symbol)
            if item.pe_ratio > 0:
                all_pe_items.append(item)
            if item.pe_ratio > 0 and item.pe_ratio < pe_threshold:
                items.append(item)
                if len(items) >= limit:
                    break
        except Exception as exc:
            print(f"[WARN] Skip {symbol}: {exc}", file=sys.stderr)

    # If threshold too strict, relax and use lowest available PE
    if not items and all_pe_items:
        print(f"[WARN] No stocks with PE < {pe_threshold}; using {len(all_pe_items)} lowest-PE available", file=sys.stderr)
        items = sorted(all_pe_items, key=lambda i: i.pe_ratio)[:limit]

    return items[:limit]


def get_stock_universe_by_mode(config: dict[str, Any], limit: int, scan_count: int, mode: str) -> list[StockItem]:
    """Route to the correct screener based on mode."""
    if mode == "losers":
        quotes = fetch_yahoo_screener("day_losers", max(scan_count, limit * 20))
        if not quotes:
            raise RuntimeError("Yahoo day_losers screener returned no results.")
        print(f"[INFO] Day losers scanned: {len(quotes)} candidates")
        return _screen_stocks_from_quotes(quotes, limit,
                                          sort_key=lambda i: i.pct_change, reverse=False)

    if mode == "low_pe":
        return _screen_low_pe(limit, scan_count)

    if mode == "trending":
        quotes = fetch_yahoo_screener("most_actives", max(scan_count, limit * 20))
        if not quotes:
            raise RuntimeError("Yahoo most_actives screener returned no results.")
        print(f"[INFO] Most actives scanned: {len(quotes)} candidates")
        return _screen_stocks_from_quotes(quotes, limit,
                                          sort_key=lambda i: i.volume, reverse=True,
                                          fetch_profile=False)

    # Default: gainers (existing behavior, respects sector/symbol P2 config)
    return get_stock_universe(config, limit, scan_count)


def domain_from_url(url: str) -> str:
    if not url:
        return ""
    parsed = urlparse(url if "://" in url else f"https://{url}")
    return parsed.netloc.replace("www.", "").strip("/")


def first_nested_value(data: Any, paths: list[list[Any]]) -> Any:
    for path in paths:
        cur = data
        for key in path:
            if isinstance(key, int):
                if not isinstance(cur, list) or len(cur) <= key:
                    cur = None
                    break
                cur = cur[key]
            elif isinstance(cur, dict):
                cur = cur.get(key)
            else:
                cur = None
                break
        if cur:
            return cur
    return ""


def compact_source_name(source: str, url: str = "") -> str:
    if source:
        return source[:28]
    domain = domain_from_url(url)
    return domain[:28] if domain else "ข่าวล่าสุด"


def news_relevance_score(item: StockItem, news: NewsItem) -> int:
    text = f"{news.title} {news.source}".lower()
    score = 0
    if item.symbol.lower() in text:
        score += 6
    for token in re.split(r"[^a-z0-9]+", item.company_name.lower()):
        if len(token) >= 4 and token in text:
            score += 4
    for token in re.split(r"[^a-z0-9]+", f"{item.industry} {item.sector}".lower()):
        if len(token) >= 5 and token in text:
            score += 2
    theme_keywords = []
    business = company_business_hint(item)
    if "ชิป" in business:
        theme_keywords = ["ai", "chip", "chips", "semiconductor", "gpu"]
    elif "ปุ๋ย" in business:
        theme_keywords = ["fertilizer", "agriculture", "crop", "farm"]
    elif "เหมือง" in business:
        theme_keywords = ["gold", "copper", "mining", "metal"]
    elif "พลังงาน" in business:
        theme_keywords = ["oil", "gas", "energy", "solar", "renewable"]
    score += sum(1 for keyword in theme_keywords if keyword in text)
    if news.image_bytes:
        score += 1
    return score


def news_overlay_text(item: StockItem, news: NewsItem) -> str:
    title = clean_text_for_font(news.title)
    lowered = title.lower()
    symbol = item.symbol.upper()
    business = company_business_hint(item)
    if symbol.lower() in lowered or any(token in lowered for token in re.split(r"[^a-z0-9]+", item.company_name.lower()) if len(token) >= 4):
        if item.pct_change < 0:
            return f"ข่าวกดดัน {symbol} ตลาดเริ่มประเมินความเสี่ยง"
        return f"ข่าวหนุน {symbol} ตลาดเริ่มประเมินมูลค่าใหม่"
    if "ชิป" in business or any(word in lowered for word in ["ai", "chip", "semiconductor", "gpu", "openai", "anthropic"]):
        return "ธีม AI ยังร้อน หุ้นชิปถูกจับตาต่อ"
    if "ปุ๋ย" in business or any(word in lowered for word in ["fertilizer", "agriculture", "crop", "farm"]):
        return "ต้นทุนเกษตรและปุ๋ยกลับมาเป็นธีมตลาด"
    if "เหมือง" in business or any(word in lowered for word in ["gold", "copper", "mining", "metal"]):
        return "สินค้าโภคภัณฑ์ขยับ หุ้นเหมืองเริ่มมีแรงเก็งกำไร"
    if "พลังงาน" in business or any(word in lowered for word in ["oil", "gas", "energy", "solar"]):
        return "ราคาพลังงานและดีมานด์เป็นตัวแปรสำคัญ"
    clean_title = re.sub(r"\s+", " ", title).strip()
    return clean_title[:72] + ("..." if len(clean_title) > 72 else "")


def news_overlay_copy(item: StockItem, news: NewsItem) -> tuple[str, str]:
    label = clean_meme_line(getattr(item, "news_card_label", ""), 28)
    detail = clean_meme_line(getattr(item, "news_card_detail", ""), 34)
    if label and detail:
        return label, detail
    return fallback_news_overlay_copy(item, news)


def fallback_news_overlay_copy(item: StockItem, news: NewsItem) -> tuple[str, str]:
    title = clean_text_for_font(news.title)
    lowered = title.lower()
    symbol = item.symbol.upper()
    business = company_business_hint(item)
    angle = company_share_angle(item)
    direct_stock = symbol.lower() in lowered or any(
        token in lowered for token in re.split(r"[^a-z0-9]+", item.company_name.lower()) if len(token) >= 4
    )

    if direct_stock and item.pct_change < 0:
        choices = [
            (f"{symbol} โดนตลาดทดสอบ", "ข่าวนี้ทำให้แรงขายน่าคิด"),
            (f"แรงขายมีที่มา", "ตลาดเริ่มจับผิดธุรกิจหลัก"),
            (f"{symbol} เสียทรงเพราะข่าว?", "วอลุ่มสูงกว่าที่ควรมองข้าม"),
        ]
        return choices[stable_index(symbol + title, len(choices))]
    if direct_stock:
        if any(word in lowered for word in ["volume", "trading", "active"]):
            return f"วอลุ่ม {symbol} ไม่ธรรมดา", "แรงซื้อขายกำลังเล่าเรื่อง"
        if any(word in lowered for word in ["earnings", "results", "profit", "revenue", "beat"]):
            return f"งบทำตลาดหันมอง {symbol}", "ตัวเลขล่าสุดเริ่มดันความมั่นใจ"
        if any(word in lowered for word in ["ipo", "deal", "acquisition", "merger"]):
            return f"{symbol} มีดีลให้ลุ้น", "ตลาดเริ่มใส่พรีเมียมข่าว"
        choices = [
            (f"{symbol} ได้แรงส่งใหม่", "ตลาดเริ่มตีความข่าวเป็นบวก"),
            (f"ข่าวนี้ปลุก {symbol}", f"{angle} ถูกดึงกลับมาโฟกัส"),
            (f"{symbol} มีสตอรี่หนุน", "แรงซื้อไม่ได้มาแบบไร้เหตุผล"),
            (f"{symbol} ถูกจับเข้าจอ", "ข่าวล่าสุดทำให้ราคาร้อนขึ้น"),
            (f"เงินเริ่มไหลหา {symbol}", "นักลงทุนกำลังซื้อ narrative"),
        ]
        return choices[stable_index(f"{symbol}:{title}:{item.pct_change}", len(choices))]
    if "ชิป" in business or any(word in lowered for word in ["ai", "chip", "semiconductor", "gpu", "openai", "anthropic"]):
        choices = [
            ("AI ยังเขย่ากระดาน", "หุ้นเทคถูกลากเข้าเกมข่าว"),
            ("ธีมชิปยังไม่จบ", "ดีมานด์ AI คือจุดที่ตลาดตาม"),
            ("ข่าว AI ทำตลาดตื่น", f"{symbol} ถูกโยงเข้ากระแสใหญ่"),
        ]
        return choices[stable_index(symbol + title, len(choices))]
    if "ปุ๋ย" in business or any(word in lowered for word in ["fertilizer", "agriculture", "crop", "farm"]):
        return "ข่าวเกษตรเริ่มมีน้ำหนัก", "ต้นทุนปุ๋ยอาจกลับมาขยับเกม"
    if "เหมือง" in business or any(word in lowered for word in ["gold", "copper", "mining", "metal"]):
        return "โลหะขยับ หุ้นเหมืองคึก", "ตลาดเริ่มเก็งรอบสินค้าโภคภัณฑ์"
    if "พลังงาน" in business or any(word in lowered for word in ["oil", "gas", "energy", "solar"]):
        return "พลังงานเป็นตัวแปรใหญ่", "ราคาขายและดีมานด์ต้องตามใกล้ๆ"
    if item.pct_change < 0:
        choices = [
            (f"{symbol} แดงแต่มีประเด็น", "ข่าวนี้อาจเป็นตัวเร่งแรงขาย"),
            (f"ตลาดเริ่มถามหาเหตุผล", f"{angle} ต้องพิสูจน์ต่อ"),
            (f"{symbol} ถูกลดความมั่นใจ", "นักลงทุนรอดูข่าวชุดถัดไป"),
        ]
        return choices[stable_index(symbol + title, len(choices))]
    choices = [
        (f"{symbol} มีเหตุให้ถูกซื้อ", "ข่าวนี้ช่วยเติมน้ำหนักฝั่งบวก"),
        ("ตลาดกำลังตีราคาใหม่", f"{angle} ถูกเล่าด้วยมุมใหม่"),
        (f"{symbol} ไม่ได้ขึ้นลอยๆ", "ข่าวล่าสุดทำให้ภาพดูน่าสนใจ"),
    ]
    return choices[stable_index(symbol + title, len(choices))]


def generate_news_overlay_copy(item: StockItem, creds: AppCreds, model: str) -> tuple[str, str] | None:
    news = next((entry for entry in item.news_items if entry.image_bytes), None)
    news = news or (item.news_items[0] if item.news_items else None)
    if not news:
        return None
    fallback = fallback_news_overlay_copy(item, news)
    if not creds.openrouter_key:
        return fallback
    prompt = (
        "คุณคือ headline editor ของเพจหุ้นไทย เขียน copy สำหรับกรอบข่าวเล็กบนรูปกราฟ\n"
        "เป้าหมาย: อ่านแล้วเข้าใจทันทีว่าข่าวเกี่ยวอะไรกับหุ้น และต้องน่าสนใจกว่า 'ข่าวหนุน/ตลาดมองโอกาส'\n"
        "ตอบแค่ 2 บรรทัด ห้ามใส่เลขข้อ ห้ามใส่คำอธิบาย ห้ามใส่ emoji\n"
        "บรรทัด 1: hook ข่าวสั้นมาก 14-28 ตัวอักษร ใส่ ticker ได้ถ้าเหมาะ\n"
        "บรรทัด 2: อธิบาย impact แบบมนุษย์อ่านรู้เรื่อง 18-34 ตัวอักษร\n"
        "หลีกเลี่ยงคำ generic: ข่าวหนุน, ข่าวกดดัน, ตลาดมองโอกาสรอบใหม่, น่าจับตา, ประเด็นข่าว\n\n"
        f"Ticker: {item.symbol}\nCompany: {item.company_name}\n"
        f"Move: {item.pct_change:+.2f}%\nVolume: {item.volume:,}\n"
        f"Business: {company_share_angle(item)}\n"
        f"Image headline: {item.headline}\n"
        f"News source: {news.source}\nNews title: {news.title}\n"
    )
    try:
        raw = chat_completion(
            api_key=creds.openrouter_key,
            base_url="https://openrouter.ai/api/v1/chat/completions",
            model=model,
            prompt=prompt,
            headers={"HTTP-Referer": "http://localhost", "X-Title": "BulkVideoCreatorApp"},
        )
        lines = [clean_meme_line(line, 38) for line in raw.splitlines() if clean_meme_line(line, 38)]
        if len(lines) >= 2:
            return clean_meme_line(lines[0], 28), clean_meme_line(lines[1], 34)
    except Exception as exc:
        print(f"[WARN] AI news card copy failed for {item.symbol}: {exc}", file=sys.stderr)
    return fallback


def format_news_for_prompt(item: StockItem) -> str:
    if not item.news_items:
        return "- No major headline found"
    lines = []
    for idx, news in enumerate(item.news_items[:3], start=1):
        source = news.source or "Unknown source"
        url = news.url or "(no URL)"
        overlay = news_overlay_text(item, news)
        lines.append(f"{idx}. Source: {source}\nTitle: {news.title}\nThai angle: {overlay}\nURL: {url}")
    return "\n".join(lines)


def append_news_sources(caption: str, item: StockItem) -> str:
    sources = [news for news in item.news_items[:3] if news.url]
    if not sources:
        return caption.strip()
    if "แหล่งข่าว" in caption and any(news.url in caption for news in sources):
        return caption.strip()
    lines = ["", "แหล่งข่าวอ้างอิง:"]
    for news in sources:
        source = news.source or domain_from_url(news.url) or "Source"
        title = clean_text_for_font(news.title)
        if len(title) > 90:
            title = title[:87].rstrip() + "..."
        lines.append(f"- {source}: {title} — {news.url}")
    return caption.strip() + "\n" + "\n".join(lines)


def stable_index(value: str, modulo: int) -> int:
    if modulo <= 0:
        return 0
    digest = hashlib.sha256(value.encode("utf-8", "ignore")).hexdigest()
    return int(digest[:8], 16) % modulo


def download_news_image(url: str) -> bytes | None:
    if not url:
        return None
    session = requests_session()
    try:
        res = session.get(url, timeout=12, allow_redirects=True)
        if res.status_code == 200 and is_supported_image_bytes(res.content):
            return res.content
    except Exception as exc:
        print(f"[WARN] News image failed: {url} ({exc})", file=sys.stderr)
    return None


def fetch_news(item: StockItem, creds: AppCreds) -> list[str]:
    headlines: list[str] = []
    news_items: list[NewsItem] = []
    session = requests_session()
    query = f"{item.symbol} OR \"{item.company_name}\""
    if creds.news_api_key:
        try:
            res = session.get(
                "https://newsapi.org/v2/everything",
                params={
                    "q": query,
                    "language": "en",
                    "sortBy": "publishedAt",
                    "pageSize": 3,
                    "apiKey": creds.news_api_key,
                },
                timeout=15,
            )
            res.raise_for_status()
            articles = res.json().get("articles", [])
            for article in articles:
                title = str(article.get("title") or "").strip()
                if not title:
                    continue
                news_items.append(
                    NewsItem(
                        title=title,
                        source=compact_source_name(str((article.get("source") or {}).get("name") or ""), str(article.get("url") or "")),
                        url=str(article.get("url") or ""),
                        image_url=str(article.get("urlToImage") or ""),
                    )
                )
                headlines.append(title)
        except Exception as exc:
            print(f"[WARN] NewsAPI failed for {item.symbol}: {exc}", file=sys.stderr)

    if len(headlines) < 3:
        try:
            import yfinance as yf

            ticker_news = getattr(yf.Ticker(item.symbol), "news", []) or []
            for entry in ticker_news:
                title = entry.get("title") or entry.get("content", {}).get("title")
                if title:
                    title_str = str(title).strip()
                    content = entry.get("content") if isinstance(entry.get("content"), dict) else {}
                    url = str(
                        entry.get("link")
                        or entry.get("url")
                        or content.get("canonicalUrl", {}).get("url")
                        or content.get("clickThroughUrl", {}).get("url")
                        or ""
                    )
                    source = str(
                        entry.get("publisher")
                        or content.get("provider", {}).get("displayName")
                        or content.get("provider", {}).get("source")
                        or ""
                    )
                    image_url = str(
                        first_nested_value(
                            entry,
                            [
                                ["thumbnail", "resolutions", 1, "url"],
                                ["thumbnail", "resolutions", 0, "url"],
                                ["content", "thumbnail", "resolutions", 1, "url"],
                                ["content", "thumbnail", "resolutions", 0, "url"],
                                ["content", "thumbnail", "originalUrl"],
                            ],
                        )
                        or ""
                    )
                    if title_str not in headlines:
                        headlines.append(title_str)
                    if not any(news.title == title_str for news in news_items):
                        news_items.append(
                            NewsItem(
                                title=title_str,
                                source=compact_source_name(source, url),
                                url=url,
                                image_url=image_url,
                            )
                        )
                if len(headlines) >= 3:
                    break
        except Exception:
            pass

    deduped_items: list[NewsItem] = []
    seen_titles: set[str] = set()
    for news in news_items:
        if not news.title or news.title in seen_titles:
            continue
        seen_titles.add(news.title)
        if news.image_url and not news.image_bytes:
            news.image_bytes = download_news_image(news.image_url)
        deduped_items.append(news)

    ranked_items = sorted(deduped_items, key=lambda news: news_relevance_score(item, news), reverse=True)[:3]
    item.news_items = ranked_items
    return [news.title for news in ranked_items] or list(dict.fromkeys([h for h in headlines if h]))[:3]


def summarize_ohlc(item: StockItem) -> str:
    rows = []
    hist = item.ohlc.tail(30)
    for idx, row in hist.iterrows():
        day = idx.strftime("%Y-%m-%d") if hasattr(idx, "strftime") else str(idx)
        rows.append(
            f"{day}: O={float(row['Open']):.2f}, H={float(row['High']):.2f}, "
            f"L={float(row['Low']):.2f}, C={float(row['Close']):.2f}, V={int(row['Volume'])}"
        )
    return "\n".join(rows)


def company_business_hint(item: StockItem) -> str:
    raw = " ".join(part for part in [item.industry, item.sector] if part).strip()
    lowered = raw.lower()
    keyword_map = [
        (("agricultural inputs", "fertilizer", "crop"), "ธุรกิจปุ๋ยและต้นทุนเกษตร"),
        (("gold", "precious metals", "mining", "metal"), "ธุรกิจเหมืองแร่"),
        (("semiconductor", "chip"), "ธุรกิจชิป"),
        (("software", "application"), "ธุรกิจซอฟต์แวร์"),
        (("internet", "content", "interactive media"), "แพลตฟอร์มออนไลน์"),
        (("biotechnology", "biotech"), "ธุรกิจไบโอเทค"),
        (("pharmaceutical", "drug"), "ธุรกิจยา"),
        (("medical", "healthcare", "health care"), "ธุรกิจสุขภาพ"),
        (("bank", "credit"), "ธุรกิจการเงิน"),
        (("insurance",), "ธุรกิจประกัน"),
        (("asset management", "capital markets"), "ธุรกิจลงทุน"),
        (("oil", "gas", "energy"), "ธุรกิจพลังงาน"),
        (("solar", "renewable"), "ธุรกิจพลังงานสะอาด"),
        (("auto", "vehicle", "ev"), "ธุรกิจยานยนต์"),
        (("retail", "consumer"), "ธุรกิจค้าปลีก"),
        (("restaurant", "food"), "ธุรกิจอาหาร"),
        (("aerospace", "defense"), "ธุรกิจอวกาศและกลาโหม"),
        (("real estate", "reit"), "ธุรกิจอสังหา"),
        (("utility", "utilities"), "ธุรกิจสาธารณูปโภค"),
    ]
    for keys, label in keyword_map:
        if any(key in lowered for key in keys):
            return label
    if item.industry:
        return item.industry[:32]
    if item.sector:
        return item.sector[:32]
    return "ธุรกิจหลัก"


def company_share_angle(item: StockItem) -> str:
    hint = company_business_hint(item)
    lowered = " ".join(part for part in [hint, item.industry, item.sector] if part).lower()
    angle_map = [
        (("ปุ๋ย", "agricultural", "fertilizer"), "ปุ๋ยและต้นทุนเกษตรที่ตลาดเริ่มมองใหม่"),
        (("เหมือง", "gold", "mining", "metal"), "ราคาสินค้าโภคภัณฑ์อาจกำลังเปลี่ยนเกม"),
        (("ชิป", "semiconductor", "chip"), "ชิปและ AI ยังเป็นธีมที่ตลาดไม่ยอมปล่อย"),
        (("ซอฟต์แวร์", "software"), "รายได้ซอฟต์แวร์กำลังถูกตีราคาใหม่"),
        (("ออนไลน์", "internet", "media"), "แพลตฟอร์มออนไลน์ยังมีแรงดึงเงินโฆษณา"),
        (("ไบโอเทค", "biotech"), "ข่าวทดลองยาอาจพลิกมูลค่าบริษัทได้เร็ว"),
        (("ยา", "pharmaceutical"), "ท่อผลิตภัณฑ์ยาคือประเด็นที่ต้องตามต่อ"),
        (("สุขภาพ", "medical", "healthcare"), "ดีมานด์สุขภาพยังเป็นเรื่องใหญ่ของตลาด"),
        (("การเงิน", "bank", "credit"), "ดอกเบี้ยและสินเชื่อกำลังชี้ทิศทางกำไร"),
        (("พลังงานสะอาด", "solar", "renewable"), "พลังงานสะอาดกลับมาเป็นธีมที่ต้องจับตา"),
        (("พลังงาน", "oil", "gas", "energy"), "ราคาพลังงานอาจเป็นตัวเร่งกำไร"),
        (("ยานยนต์", "auto", "vehicle", "ev"), "ยอดขายและมาร์จิ้นคือจุดชี้ชะตาราคา"),
        (("ค้าปลีก", "retail"), "กำลังซื้อผู้บริโภคคือคำตอบของรอบนี้"),
    ]
    for keys, angle in angle_map:
        if any(key in lowered for key in keys):
            return angle
    return f"{hint}ที่ตลาดอาจกำลังตีราคาใหม่"


def headline_is_too_generic(text: str) -> bool:
    lowered = clean_text_for_font(text).lower()
    bad_phrases = [
        "เกิดอะไรขึ้น",
        "เกิดอะไรขึ้นกับ",
        "หุ้นตัวนี้",
        "ดาวเด่นวันนี้",
        "ปิดตลาดสวย",
        "แรงซื้อดันราคา",
        "โดดเด่น",
        "ผู้นำนวัตกรรม",
        "น่าจับตา",
        "ต้องจับตา",
        "ไม่ควรมองข้าม",
        "กำลังถูกจับตา",
    ]
    return any(phrase in lowered for phrase in bad_phrases)


def headline_variant_seed(item: StockItem, mode: str, salt: str = "") -> str:
    return (
        f"{mode}:{getattr(item, 'symbol', '')}:"
        f"{getattr(item, 'company_name', '')}:"
        f"{getattr(item, 'pct_change', 0):.2f}:"
        f"{getattr(item, 'volume', 0)}:"
        f"{getattr(item, 'overlay_variant', '')}:{salt}"
    )


def select_headline_variant(choices: list[str], item: StockItem, mode: str, salt: str) -> str:
    if not choices:
        return ""
    try:
        variant = int(getattr(item, "overlay_variant", 0) or 0)
    except (TypeError, ValueError):
        variant = 0
    if salt == "hook" and variant > 0:
        return choices[(variant - 1) % len(choices)]
    return choices[stable_index(headline_variant_seed(item, mode, salt), len(choices))]


def viral_hook_choices(item: StockItem, mode: str) -> list[str]:
    symbol = item.symbol.upper()
    business = company_business_hint(item)
    if len(business) > 18:
        business = business[:18].rstrip()
    gain_hooks = [
        f"{symbol} รอบนี้มีอะไรซ่อนอยู่?",
        f"ทำไมเงินไหลเข้า {symbol}?",
        f"{symbol} กำลังส่งสัญญาณอะไร?",
        "ตลาดเห็นอะไรที่คนยังไม่เห็น?",
        "แรงซื้อรอบนี้ธรรมดาหรือเปล่า?",
        f"{symbol} กำลังถูกตีราคาใหม่?",
        "จังหวะนี้มีข่าวดีแค่ไหน?",
        "ราคาเด้งเพราะอะไร?",
        "รอบนี้แค่เก็งหรือมีของจริง?",
        "นักลงทุนกำลังซื้อเรื่องอะไร?",
        "ทำไมวอลุ่มถึงมาพร้อมราคา?",
        "แรงซื้อกลับมาเพราะอะไร?",
        f"{symbol} มีสตอรี่ใหม่หรือยัง?",
        "ตลาดกำลังมองข้ามอะไร?",
        "ดีลนี้เริ่มน่าสนใจตรงไหน?",
        "เงินร้อนกำลังเข้าเกม?",
        "ราคาเริ่มนำข่าวไปก่อน?",
        "รอบนี้ตลาดเชื่ออะไร?",
        "สัญญาณบวกเริ่มชัดหรือยัง?",
        "กำลังเกิด re-rating หรือเปล่า?",
        "ทำไมวันนี้มีแรงไล่ราคา?",
        "ข่าวดีพอเปลี่ยนเกมไหม?",
        "โมเมนตัมรอบนี้มาจากไหน?",
        "ตลาดเริ่มให้พรีเมียมแล้ว?",
        "แรงซื้อแบบนี้ต้องอ่านอะไร?",
        f"{business} กลับมาเป็นธีมหรือยัง?",
        "กำไรหรือข่าวที่ตลาดกำลังซื้อ?",
        "ราคาเขียวเพราะคนรู้อะไร?",
        "รอบนี้ไม่ใช่แค่เด้ง?",
        "ตลาดกำลังโหวตด้วยเงิน?",
        "หุ้นนี้เริ่มหลุดเรดาร์เดิม?",
        "ทำไมกราฟเริ่มตื่น?",
        "แรงเก็งกำไรกำลังจุดติด?",
        "มี catalyst ใหม่หรือเปล่า?",
        "ตลาดกำลังเปลี่ยนมุมมอง?",
        "วันนี้เงินเลือกข้างไหน?",
        "สตอรี่บวกเริ่มชนะความกังวล?",
        "ราคาเริ่มสะท้อนข่าวดี?",
        "นักลงทุนเริ่มกลับมาเชื่อ?",
        "หุ้นนี้กำลังเข้าโหมดร้อน?",
        "กระแสซื้อเริ่มจริงจัง?",
        "ตลาดกำลังไล่ narrative?",
        "รอบนี้ใครกำลังเก็บของ?",
        "สัญญาณนี้ควรอ่านให้ทัน?",
        "แรงซื้อกำลังบอกอะไร?",
        "ราคาวิ่งก่อนข่าวใหญ่?",
        "วันนี้มีอะไรเปลี่ยนในมุมตลาด?",
        "หุ้นนี้กำลังถูกค้นพบใหม่?",
        "โมเมนตัมกำลังเปลี่ยนมือ?",
        "ตลาดเริ่มมองอนาคตใหม่?",
    ]
    loser_hooks = [
        f"{symbol} แดงแรงเพราะอะไร?",
        "แรงขายรอบนี้บอกอะไร?",
        "ตลาดกำลังกลัวเรื่องไหน?",
        "ราคาลงแต่ประเด็นยังไม่จบ",
        "นี่คือ panic หรือโอกาส?",
        "แรงขายเริ่มเกินจริงไหม?",
        f"{symbol} ถูกทดสอบรอบใหม่",
        "ตลาดกำลังตัดสินข่าวนี้?",
        "ความเสี่ยงไหนกำลังกดราคา?",
        "จุดเปลี่ยนอยู่ตรงไหน?",
        "แดงวันนี้ต้องอ่านอะไร?",
        "นักลงทุนกำลังหนีอะไร?",
        "ข่าวร้ายสะท้อนหมดหรือยัง?",
        "ราคาลงแต่เกมยังไม่จบ?",
        "แรงขายนี้มีเหตุผลแค่ไหน?",
    ]
    low_pe_hooks = [
        "ถูกจริงหรือแค่ดูถูก?",
        "หุ้น Value รอบนี้มีอะไร?",
        "P/E ต่ำพอให้ตลาดหันมอง?",
        "ราคานี้ซ่อนมูลค่าไหม?",
        "ตลาดมองข้ามของดีหรือเปล่า?",
        "ถูกแต่มี trigger หรือยัง?",
        "มูลค่ากับราคากำลังสวนกัน?",
        "Value play เริ่มกลับมา?",
        f"{symbol} ถูกเกินไปไหม?",
        "ดีลนี้สายพื้นฐานต้องอ่าน",
    ]
    trending_hooks = [
        "วอลุ่มพุ่งเพราะใครกำลังเล่น?",
        "ตลาดแห่เข้าหาหุ้นนี้ทำไม?",
        "หุ้นนี้กำลังขึ้นหน้าจอใหญ่?",
        "ปริมาณซื้อขายบอกอะไร?",
        "คนทั้งตลาดกำลังมองอะไร?",
        "ข่าวไหนทำให้วอลุ่มเดือด?",
        "ทำไมวันนี้ซื้อขายผิดปกติ?",
        "กระแสตลาดกำลังไหลไปไหน?",
        "หุ้นนี้กำลังเป็นประเด็นร้อน?",
        "วอลุ่มมาแบบนี้ต้องมีเรื่อง",
    ]
    if mode == "losers" or item.pct_change < 0:
        return loser_hooks + gain_hooks
    if mode == "low_pe":
        return low_pe_hooks + gain_hooks
    if mode == "trending":
        return trending_hooks + gain_hooks
    return gain_hooks


def viral_second_line_choices(item: StockItem, mode: str, change: str) -> list[str]:
    symbol = item.symbol.upper()
    if mode == "losers" or item.pct_change < 0:
        return [
            f"**{symbol}** หลุด {change} ตลาดเริ่มกังวล",
            f"**{symbol}** ร่วง {change} แต่วอลุ่มยังแรง",
            f"**{symbol}** แดง {change} นักลงทุนเริ่มชั่งใจ",
            f"**{symbol}** ถูกขาย {change} ประเด็นเริ่มหนัก",
            f"**{symbol}** ดิ่ง {change} ตลาดรอคำตอบ",
        ]
    if mode == "low_pe":
        pe = f"P/E {item.pe_ratio:.1f}x" if item.pe_ratio > 0 else "Valuation ต่ำ"
        return [
            f"**{symbol}** {pe} แต่เริ่มมีแรงขยับ",
            f"**{symbol}** ราคายังถูก แต่ตลาดเริ่มสนใจ",
            f"**{symbol}** พื้นฐานเริ่มถูกหยิบมาคุย",
            f"**{symbol}** มูลค่าถูก แต่ต้องมี trigger",
            f"**{symbol}** สาย Value เริ่มเปิดจอ",
        ]
    if mode == "trending":
        return [
            f"**{symbol}** วอลุ่มหนาแน่น ตลาดเริ่มถามหาเหตุผล",
            f"**{symbol}** ซื้อขายคึกคักจนต้องดูข่าว",
            f"**{symbol}** ติดเรดาร์ด้วยวอลุ่มผิดปกติ",
            f"**{symbol}** คนแห่เทรดจนกระแสเริ่มมา",
            f"**{symbol}** วอลุ่มเดือด ราคากำลังส่งสัญญาณ",
        ]
    return [
        f"**{symbol}** ทะยาน {change} แรงซื้อกลับมา",
        f"**{symbol}** พุ่ง {change} วอลุ่มหนาแน่น",
        f"**{symbol}** วิ่ง {change} ตลาดเริ่มตื่น",
        f"**{symbol}** บวก {change} กระแสซื้อชัด",
        f"**{symbol}** เด้งแรง {change} ไม่ใช่ภาพนิ่ง",
        f"**{symbol}** ราคาพุ่ง {change} แรงเก็งกำไรมา",
        f"**{symbol}** ปิดบวก {change} พร้อมสตอรี่หนุน",
        f"**{symbol}** ขยับ {change} นักลงทุนกลับมามอง",
    ]


def viral_third_line_choices(item: StockItem, angle: str) -> list[str]:
    business = company_business_hint(item)
    if len(business) > 22:
        business = business[:22].rstrip()
    choices = [
        angle,
        f"**{business}** กำลังถูกตีราคาใหม่",
        f"**{business}** เริ่มกลับมาอยู่ในเรดาร์",
        f"**{business}** คือธีมที่ตลาดหยิบมาคุย",
        f"**{business}** อาจเป็นตัวเร่งรอบนี้",
        f"**{business}** กำลังเปลี่ยนมุมมองนักลงทุน",
        f"**{business}** มีข่าวหนุนให้ตลาดคิดต่อ",
        f"**{business}** ยังมีประเด็นให้ตามต่อ",
        f"**{business}** กำลังเจอแรงประเมินมูลค่าใหม่",
        f"**{business}** อาจไม่ใช่เรื่องเล็กของรอบนี้",
    ]
    return choices


def build_viral_headline(item: StockItem, mode: str) -> str:
    change_abs = abs(item.pct_change)
    change = f"+{change_abs:.2f}%" if item.pct_change >= 0 else f"-{change_abs:.2f}%"
    angle = company_share_angle(item)
    lines = [
        select_headline_variant(viral_hook_choices(item, mode), item, mode, "hook"),
        select_headline_variant(viral_second_line_choices(item, mode, change), item, mode, "line2"),
        select_headline_variant(viral_third_line_choices(item, angle), item, mode, "line3"),
    ]
    return "\n".join(lines)


def polish_image_headline(item: StockItem, mode: str) -> str:
    lines = [line.strip() for line in str(item.headline or "").splitlines() if line.strip()]
    if len(lines) != 3 or headline_is_too_generic(item.headline):
        return build_viral_headline(item, mode)
    # Keep useful AI-written facts, but control the opening hook so batches do not
    # all start with the same stock-page cliché.
    lines[0] = select_headline_variant(viral_hook_choices(item, mode), item, mode, "hook")
    return "\n".join(lines[:3])


def generate_caption(item: StockItem, creds: AppCreds, model: str, mode: str = "gainers") -> str:
    news_block = format_news_for_prompt(item)
    business_hint = company_business_hint(item)
    mode_instruction = {
        "gainers": "วิเคราะห์ว่าทำไมหุ้นตัวนี้ราคาพุ่งขึ้นสูงวันนี้ เน้นโมเมนตัมและแรงซื้อ",
        "losers":  "วิเคราะห์ว่าทำไมหุ้นตัวนี้ราคาร่วงลงมา บอกว่าผู้ลงทุนควรระวังหรือน่าซื้อที่จุดนี้หรือไม่",
        "low_pe":  f"เน้นวิเคราะห์ว่าหุ้นตัวนี้มี P/E ต่ำ ({item.pe_ratio:.1f}x) และน่าสนใจในเชิง Value Investing อย่างไร",
        "trending": "วิเคราะห์ว่าทำไมหุ้นตัวนี้ถึงมีปริมาณซื้อขายสูงมาก และมีข่าวอะไรน่าจับตาวันนี้",
    }.get(mode, "วิเคราะห์ว่าทำไมหุ้นตัวนี้ถึงน่าสนใจวันนี้")

    pe_line = f"P/E Ratio: {item.pe_ratio:.2f}x\n" if item.pe_ratio > 0 else ""
    prompt = (
        f"Analyze this stock. {mode_instruction}\n"
        "Write a short, engaging 2-3 sentence summary in Thai. Tone: Professional, exciting, "
        "and highly engaging for a financial Facebook page.\n\n"
        f"Symbol: {item.symbol}\nCompany: {item.company_name}\n"
        f"Sector: {item.sector or 'Unknown'}\nIndustry: {item.industry or 'Unknown'}\n"
        f"Business hint in Thai: {business_hint}\n"
        f"Price: {item.price:.2f}\n% Change: {item.pct_change:.2f}%\nVolume: {item.volume:,}\n"
        f"{pe_line}"
        f"OHLC last 30 trading days:\n{summarize_ohlc(item)}\n\nNews:\n{news_block}\n\n"
        "Return your response EXACTLY in this format:\n\n"
        "IMAGE_TEXT:\n"
        "[Write exactly 3 complete Thai headline lines for a viral Facebook image.\n"
        "Separate each line with a newline character. Never break a word across lines.\n"
        "Line 1: Curiosity hook that makes readers ask 'why', 16-30 Thai characters. Example: ตลาดกำลังตีราคาใหม่?\n"
        "Line 2: Main fact with tension, 24-48 Thai characters, include **stock_symbol** and the % move. Example: **NVDA** พุ่งแรง ไม่ใช่แค่เด้งธรรมดา\n"
        "Line 3: Smart business angle, 28-52 Thai characters. Example: **ชิปและ AI** ยังเป็นธีมที่ตลาดไม่ยอมปล่อย\n"
        "Use **double asterisks** around ONLY the stock symbol and the business/activity phrase on Line 3.\n"
        "NO EMOJIS. Do NOT start Line 1 with เกิดอะไรขึ้นกับ or เกิดอะไรขึ้น. Avoid generic phrases like หุ้นตัวนี้, น่าจับตา, ดาวเด่นวันนี้, ปิดตลาดสวย, แรงซื้อดันราคา, ไม่ควรมองข้าม.\n"
        "Make the 3 lines feel like one shareable story, not 3 isolated labels.]\n\n"
        "POST_CAPTION:\n"
        "[Write a concise but professional stock analysis article in Thai for a financial Facebook page.\n"
        "Do NOT dump raw English news titles. Summarize what the news means for this stock in clear Thai.\n"
        "Make the opening interesting, specific, and shareable. Be concise and useful.\n"
        "Use this structure:\n"
        "1. 🧠 Hook headline — one line summarizing the stock's story today\n"
        "2. Price summary — current price, % change, volume, 1-month trend in 2-3 sentences\n"
        "3. 🔑 News angle — 2-3 bullets explaining the key news/theme and why it may affect investor sentiment\n"
        "4. 📍 Key levels — Support and Resistance zones from the chart (2-3 levels each)\n"
        "5. ⚠️ Risks — 2-3 key risks investors should watch\n"
        "6. 🎯 Action plan — Brief recommendation for different investor types\n"
        "7. แหล่งข่าวอ้างอิง — include source names and URLs from the News block\n"
        "8. Disclaimer line\n\n"
        "Keep it concise (180-320 words). Use Thai mixed with English financial terms naturally.\n"
        "Use emojis sparingly for section headers only. Make it feel professional like a real analyst wrote it.]\n\n"
        "DO NOT use any emojis in IMAGE_TEXT."
    )

    if creds.openrouter_key:
        try:
            return chat_completion(
                api_key=creds.openrouter_key,
                base_url="https://openrouter.ai/api/v1/chat/completions",
                model=model,
                prompt=prompt,
                headers={"HTTP-Referer": "http://localhost", "X-Title": "BulkVideoCreatorApp"},
            )
        except Exception as exc:
            print(f"[WARN] OpenRouter caption failed for {item.symbol}: {exc}", file=sys.stderr)

    if creds.openai_key:
        try:
            return chat_completion(
                api_key=creds.openai_key,
                base_url="https://api.openai.com/v1/chat/completions",
                model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
                prompt=prompt,
                headers={},
            )
        except Exception as exc:
            print(f"[WARN] OpenAI caption failed for {item.symbol}: {exc}", file=sys.stderr)

    headline_str = item.news[0] if item.news else "ตลาดกำลังจับตาหุ้นตัวนี้"
    if mode == "losers":
        return (
            f"IMAGE_TEXT:\nหุ้นร่วงแรงต้องจับตา\n**{item.symbol}** ดิ่ง {abs(item.pct_change):.2f}% วันนี้\n{business_hint} ต้องอ่านก่อนตัดสินใจ\n\n"
            f"POST_CAPTION:\n{item.symbol} วันนี้ราคาร่วงลง {abs(item.pct_change):.2f}% "
            f"มาอยู่ที่ {item.price:.2f} ดอลลาร์ ด้วยวอลุ่ม {item.volume:,} หุ้น\n\n"
            f"ข่าวสำคัญ: \"{headline_str}\"\n\nอาจเป็นปัจจัยที่ส่งผลกระทบ นักลงทุนควรติดตามสถานการณ์อย่างใกล้ชิด"
        )
    if mode == "low_pe":
        pe_str = f"P/E {item.pe_ratio:.1f}x" if item.pe_ratio > 0 else "P/E ต่ำ"
        return (
            f"IMAGE_TEXT:\nหุ้นคุณค่าน่าสะสม\n**{item.symbol}** เด่นด้วย {pe_str}\n{business_hint} ที่ไม่ควรมองข้าม\n\n"
            f"POST_CAPTION:\n{item.symbol} น่าสนใจในเชิง Value Investing ด้วย{pe_str} "
            f"ราคาปัจจุบัน {item.price:.2f} ดอลลาร์ เปลี่ยนแปลง {item.pct_change:+.2f}%\n\nข่าว \"{headline_str}\" เป็นปัจจัยที่ควรติดตาม"
        )
    if mode == "trending":
        return (
            f"IMAGE_TEXT:\nวอลุ่มซื้อขายเดือด\n**{item.symbol}** ติดเทรนด์ตลาดวันนี้\n{business_hint} กำลังถูกจับตา\n\n"
            f"POST_CAPTION:\n{item.symbol} มีปริมาณซื้อขายพุ่งสูงถึง {item.volume:,} หุ้นวันนี้ "
            f"ราคาเปลี่ยนแปลง {item.pct_change:+.2f}% มาอยู่ที่ {item.price:.2f} ดอลลาร์\n\nข่าว \"{headline_str}\" ทำให้เป็นหุ้นที่ต้องจับตาเป็นพิเศษ"
        )
    return (
        f"IMAGE_TEXT:\nหุ้นพุ่งแรงประจำวัน\n**{item.symbol}** บวก {item.pct_change:.2f}% นักลงทุนจับตา\n{business_hint} อาจเป็นตัวเร่งเกม\n\n"
        f"POST_CAPTION:\n{item.symbol} วันนี้โดดเด่นด้วยแรงซื้อหนุนให้ราคาปรับขึ้น {item.pct_change:.2f}% "
        f"มาปิดใกล้ {item.price:.2f} ดอลลาร์ พร้อมวอลุ่ม {item.volume:,} หุ้น\n\nข่าวล่าสุดอย่าง \"{headline_str}\" ทำให้หุ้นตัวนี้น่าจับตาสำหรับเพจการเงินวันนี้"
    )


def chat_completion(api_key: str, base_url: str, model: str, prompt: str, headers: dict[str, str]) -> str:
    session = requests_session()
    res = session.post(
        base_url,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            **headers,
        },
        json={
            "model": model,
            "messages": [
                {"role": "system", "content": "You write concise, accurate Thai financial social captions."},
                {"role": "user", "content": prompt},
            ],
            "temperature": 0.75,
            "max_tokens": 900,
        },
        timeout=45,
    )
    res.raise_for_status()
    data = res.json()
    return data["choices"][0]["message"]["content"].strip()


def is_supported_image_bytes(image_bytes: bytes) -> bool:
    if not image_bytes or len(image_bytes) < 128:
        return False
    try:
        from PIL import Image

        with Image.open(io.BytesIO(image_bytes)) as img:
            img.verify()
        return True
    except Exception:
        return False


def logo_candidate_urls(item: StockItem) -> list[str]:
    symbol = quote(item.symbol.upper())
    urls = [
        f"https://financialmodelingprep.com/image-stock/{symbol}.png",
        f"https://storage.googleapis.com/iex/api/logos/{symbol}.png",
    ]
    if item.domain:
        domain = item.domain.strip("/")
        urls.extend(
            [
                f"https://logo.clearbit.com/{domain}",
                f"https://www.google.com/s2/favicons?domain={domain}&sz=256",
                f"https://icons.duckduckgo.com/ip3/{domain}.ico",
                f"https://{domain}/favicon.ico",
                f"https://{domain}/apple-touch-icon.png",
                f"https://{domain}/apple-touch-icon-precomposed.png",
                f"https://{domain}/favicon-96x96.png",
            ]
        )
    return list(dict.fromkeys(urls))


def fetch_logo(item: StockItem) -> bytes | None:
    session = requests_session()
    for url in logo_candidate_urls(item):
        try:
            res = session.get(url, timeout=12, allow_redirects=True)
            if res.status_code == 200 and is_supported_image_bytes(res.content):
                print(f"[INFO] Logo found for {item.symbol}: {url}")
                return res.content
        except Exception as exc:
            print(f"[WARN] Logo candidate failed for {item.symbol}: {url} ({exc})", file=sys.stderr)
    print(f"[WARN] Logo not found for {item.symbol}; using ticker fallback.", file=sys.stderr)
    return None


def remove_logo_background(image_bytes: bytes) -> bytes:
    try:
        import importlib.util

        if importlib.util.find_spec("onnxruntime") is None:
            return image_bytes
        from rembg import remove

        return remove(image_bytes)
    except BaseException:
        return image_bytes


def generate_chart(item: StockItem, output_path: Path) -> Path:
    import mplfinance as mpf

    hist = item.ohlc.copy()
    if "Adj Close" in hist.columns:
        hist = hist.drop(columns=["Adj Close"])
    market_colors = mpf.make_marketcolors(up=BULL, down=BEAR, edge="inherit", wick="inherit", volume=BULL)
    style = mpf.make_mpf_style(
        base_mpf_style="nightclouds",
        marketcolors=market_colors,
        facecolor=BG,
        figcolor=BG,
        gridcolor=BG,
        gridstyle="",
        rc={"axes.labelcolor": MUTED, "xtick.color": MUTED, "ytick.color": MUTED},
    )
    mpf.plot(
        hist.tail(30),
        type="candle",
        volume=False,
        style=style,
        axisoff=True,
        tight_layout=True,
        figsize=(9.6, 5.8),
        savefig=dict(fname=str(output_path), dpi=160, bbox_inches="tight", pad_inches=0.03, facecolor=BG),
    )
    return output_path


def load_font(name: str, size: int):
    from PIL import ImageFont

    path = FONT_DIR / name
    if path.exists():
        return ImageFont.truetype(str(path), size=size)
    return ImageFont.load_default()


def font_path_for_name(font_name: str) -> Path | None:
    path = FONT_DIR / font_name
    return path if path.exists() else None


def shaped_text_available() -> bool:
    try:
        import freetype  # noqa: F401
        import uharfbuzz  # noqa: F401

        return True
    except Exception:
        return False


@functools.lru_cache(maxsize=2048)
def shaped_text_layout(font_name: str, size: int, text: str) -> tuple[float, tuple[float, float, float, float], tuple[tuple[int, float, float, float], ...]]:
    import freetype
    import uharfbuzz as hb

    path = font_path_for_name(font_name)
    if not path:
        return 0, (0, 0, 0, 0), ()
    font_bytes = path.read_bytes()
    hb_face = hb.Face(font_bytes)
    hb_font = hb.Font(hb_face)
    hb_font.scale = (size * 64, size * 64)
    buf = hb.Buffer()
    buf.add_str(text)
    buf.guess_segment_properties()
    hb.shape(hb_font, buf)

    ft_face = freetype.Face(str(path))
    ft_face.set_pixel_sizes(0, size)
    pen_x = 0.0
    glyphs: list[tuple[int, float, float, float]] = []
    min_x = min_y = 0.0
    max_x = max_y = 0.0
    first_box = True

    for info, pos in zip(buf.glyph_infos, buf.glyph_positions):
        gid = int(info.codepoint)
        x_offset = pos.x_offset / 64.0
        y_offset = pos.y_offset / 64.0
        x_advance = pos.x_advance / 64.0
        ft_face.load_glyph(gid, freetype.FT_LOAD_DEFAULT)
        glyph = ft_face.glyph
        glyph.render(freetype.FT_RENDER_MODE_NORMAL)
        bitmap = glyph.bitmap
        x0 = pen_x + x_offset + glyph.bitmap_left
        y0 = -(y_offset + glyph.bitmap_top)
        x1 = x0 + bitmap.width
        y1 = y0 + bitmap.rows
        if bitmap.width and bitmap.rows:
            if first_box:
                min_x, min_y, max_x, max_y = x0, y0, x1, y1
                first_box = False
            else:
                min_x = min(min_x, x0)
                min_y = min(min_y, y0)
                max_x = max(max_x, x1)
                max_y = max(max_y, y1)
        glyphs.append((gid, x_offset, y_offset, x_advance))
        pen_x += x_advance

    if first_box:
        min_x, min_y, max_x, max_y = 0.0, -size * 0.75, max(pen_x, 1.0), size * 0.25
    return pen_x, (min_x, min_y, max_x, max_y), tuple(glyphs)


def shaped_text_width(font_name: str, size: int, text: str) -> float:
    if not shaped_text_available():
        return 0.0
    width, _, _ = shaped_text_layout(font_name, size, clean_text_for_font_preserve_spaces(text))
    return width


def _rgba(fill) -> tuple[int, int, int, int]:
    from PIL import ImageColor

    if isinstance(fill, tuple):
        if len(fill) == 4:
            return fill
        if len(fill) == 3:
            return fill + (255,)
    rgb = ImageColor.getrgb(str(fill))
    if len(rgb) == 4:
        return rgb
    return rgb + (255,)


def _bitmap_to_mask(bitmap):
    from PIL import Image

    width, rows = bitmap.width, bitmap.rows
    if width <= 0 or rows <= 0:
        return None
    raw = bytes(bitmap.buffer)
    pitch = abs(bitmap.pitch) or width
    if pitch != width:
        raw = b"".join(raw[row * pitch: row * pitch + width] for row in range(rows))
    return Image.frombytes("L", (width, rows), raw)


def draw_shaped_text(image, xy: tuple[int, int], text: str, font_name: str, size: int, fill, stroke_width: int = 0, stroke_fill="#000000") -> None:
    import freetype
    from PIL import Image, ImageDraw

    if not shaped_text_available():
        draw = ImageDraw.Draw(image)
        draw.text(xy, text, font=load_font(font_name, size), fill=fill, stroke_width=stroke_width, stroke_fill=stroke_fill)
        return
    path = font_path_for_name(font_name)
    if not path:
        return
    text = clean_text_for_font_preserve_spaces(text)
    _, bbox, glyphs = shaped_text_layout(font_name, size, text)
    if not glyphs:
        return

    ft_face = freetype.Face(str(path))
    ft_face.set_pixel_sizes(0, size)
    base_x = float(xy[0])
    baseline_y = float(xy[1]) - bbox[1]

    def paste_run(dx: int, dy: int, color) -> None:
        pen_x = base_x
        rgba = _rgba(color)
        for gid, x_offset, y_offset, x_advance in glyphs:
            ft_face.load_glyph(gid, freetype.FT_LOAD_DEFAULT)
            glyph = ft_face.glyph
            glyph.render(freetype.FT_RENDER_MODE_NORMAL)
            bitmap = glyph.bitmap
            mask = _bitmap_to_mask(bitmap)
            if mask is not None:
                gx = int(round(pen_x + x_offset + glyph.bitmap_left + dx))
                gy = int(round(baseline_y - (y_offset + glyph.bitmap_top) + dy))
                glyph_img = Image.new("RGBA", mask.size, rgba)
                image.paste(glyph_img, (gx, gy), mask)
            pen_x += x_advance

    if stroke_width > 0:
        offsets = [
            (dx, dy)
            for dx in range(-stroke_width, stroke_width + 1)
            for dy in range(-stroke_width, stroke_width + 1)
            if dx * dx + dy * dy <= stroke_width * stroke_width and (dx or dy)
        ]
        for dx, dy in offsets:
            paste_run(dx, dy, stroke_fill)
    paste_run(0, 0, fill)


def fit_font_to_width(draw, text: str, font_name: str, base_size: int, min_size: int, max_width: int):
    size = base_size
    while size > min_size:
        font = load_font(font_name, size)
        if draw.textlength(text, font=font) <= max_width:
            return font
        size -= 1
    return load_font(font_name, min_size)


def ellipsize_to_width(draw, text: str, font, max_width: int) -> str:
    if draw.textlength(text, font=font) <= max_width:
        return text
    suffix = "..."
    trimmed = text
    while trimmed and draw.textlength(trimmed + suffix, font=font) > max_width:
        trimmed = trimmed[:-1]
    return (trimmed.rstrip() + suffix) if trimmed else suffix


def image_from_logo_bytes(image_bytes: bytes | None, symbol: str):
    from PIL import Image, ImageDraw

    if image_bytes:
        try:
            logo = Image.open(io.BytesIO(remove_logo_background(image_bytes))).convert("RGBA")
            logo.thumbnail((120, 120), Image.Resampling.LANCZOS)
            canvas = Image.new("RGBA", (120, 120), (0, 0, 0, 0))
            canvas.alpha_composite(logo, ((120 - logo.width) // 2, (120 - logo.height) // 2))
            return canvas
        except Exception:
            pass

    fallback = Image.new("RGBA", (120, 120), (22, 27, 34, 255))
    draw = ImageDraw.Draw(fallback)
    draw.ellipse((4, 4, 116, 116), fill=(13, 17, 23, 255), outline=(63, 185, 80, 255), width=4)
    font = load_font("Kanit-Bold.ttf", 34)
    text = symbol[:3]
    bbox = draw.textbbox((0, 0), text, font=font)
    draw.text(((120 - (bbox[2] - bbox[0])) / 2, (120 - (bbox[3] - bbox[1])) / 2 - 3), text, font=font, fill=TEXT)
    return fallback


def draw_wrapped_text(draw, text: str, xy: tuple[int, int], font, fill: str, max_width: int, line_gap: int = 8) -> int:
    x, y = xy
    lines: list[str] = []
    for paragraph in text.splitlines():
        paragraph = paragraph.strip()
        if not paragraph:
            lines.append("")
            continue
        words = paragraph.split()
        if len(words) == 1:
            lines.extend(wrap_long_token(draw, paragraph, font, max_width))
            continue
        current = ""
        for word in words:
            candidate = f"{current} {word}".strip()
            if draw.textlength(candidate, font=font) <= max_width:
                current = candidate
            else:
                if current:
                    lines.append(current)
                if draw.textlength(word, font=font) > max_width:
                    split = wrap_long_token(draw, word, font, max_width)
                    lines.extend(split[:-1])
                    current = split[-1]
                else:
                    current = word
        if current:
            lines.append(current)

    for line in lines:
        draw.text((x, y), line, font=font, fill=fill)
        bbox = draw.textbbox((x, y), line or " ", font=font)
        y += bbox[3] - bbox[1] + line_gap
    return y


def wrap_long_token(draw, token: str, font, max_width: int) -> list[str]:
    lines: list[str] = []
    current = ""
    for ch in token:
        candidate = current + ch
        if draw.textlength(candidate, font=font) <= max_width or not current:
            current = candidate
        else:
            lines.append(current)
            current = ch
    if current:
        lines.append(current)
    return lines or [token]


THAI_COMBINING_ORDER = {
    **{chr(code): 1 for code in range(0x0E38, 0x0E3B)},  # lower vowels
    "\u0e31": 2,
    **{chr(code): 2 for code in range(0x0E34, 0x0E38)},  # upper vowels
    "\u0e47": 2,
    **{chr(code): 3 for code in range(0x0E48, 0x0E4C)},  # tone marks
    "\u0e4d": 4,
    "\u0e4e": 4,
}


def normalize_thai_marks(text: str) -> str:
    chars = list(str(text or ""))
    result: list[str] = []
    i = 0
    while i < len(chars):
        ch = chars[i]
        result.append(ch)
        i += 1
        marks: list[str] = []
        while i < len(chars) and chars[i] in THAI_COMBINING_ORDER:
            marks.append(chars[i])
            i += 1
        if marks:
            result.extend(sorted(marks, key=lambda mark: THAI_COMBINING_ORDER.get(mark, 9)))
    return "".join(result)


def clean_text_for_font(text: str) -> str:
    # Remove emojis and unrenderable characters for basic Thai/English fonts
    import re
    cleaned = re.sub(r'[^\u0E00-\u0E7F\u0020-\u007E\u2013\u2014\u2018\u2019\u201C\u201D\*]', '', text)
    return normalize_thai_marks(cleaned).strip()


def clean_text_for_font_preserve_spaces(text: str) -> str:
    import re
    cleaned = re.sub(r'[^\u0E00-\u0E7F\u0020-\u007E\u2013\u2014\u2018\u2019\u201C\u201D\*]', '', str(text or ""))
    return normalize_thai_marks(cleaned)


def draw_rich_text_wrapped(draw, text: str, xy: tuple[int, int], font, default_color: str, highlight_color: str, max_width: int, line_gap: int = 8) -> int:
    """Render a 3-line headline with per-line auto-scaling and tight highlights.

    Each line from the AI is separated by \\n.  For each line we measure the
    total width at the base font size, and if it exceeds *max_width* the font
    is scaled down so it fits perfectly – no mid-word wrapping.

    Highlighted text (wrapped in **) gets a tight background pill matching
    the reference design.
    """
    import re
    x_start, y = xy
    text = clean_text_for_font(text)

    # ── Split into explicit lines (AI outputs \n between them) ──
    raw_lines = [ln.strip() for ln in text.split('\n') if ln.strip()]
    if not raw_lines:
        return y

    # Base font info
    base_font_name = getattr(font, 'path', None) or "Kanit-Bold.ttf"
    # Try to get the base font path; fall back gracefully
    try:
        base_size = font.size
    except AttributeError:
        base_size = 60

    pad_x, pad_y = 4, 2  # Tight highlight padding like reference image

    for raw_line in raw_lines:
        # Parse **highlights** within this line
        segments: list[tuple[str, bool]] = []
        for part in re.split(r'(\*\*.*?\*\*)', raw_line):
            if not part:
                continue
            if part.startswith('**') and part.endswith('**'):
                segments.append((part[2:-2], True))
            else:
                segments.append((part, False))

        plain_text = ''.join(seg[0] for seg in segments)
        if not plain_text.strip():
            y += line_gap
            continue

        # ── Auto-scale: find the largest font size that fits ──
        cur_size = base_size
        cur_font = load_font(base_font_name, cur_size)
        total_w = draw.textlength(plain_text, font=cur_font)
        while total_w > max_width and cur_size > 24:
            cur_size -= 2
            cur_font = load_font(base_font_name, cur_size)
            total_w = draw.textlength(plain_text, font=cur_font)

        ref_bbox = draw.textbbox((0, 0), "Aygjป์", font=cur_font)
        line_h = ref_bbox[3] - ref_bbox[1]

        # ── 1st pass: draw highlight backgrounds ──
        rx = x_start
        for seg_text, is_hl in segments:
            seg_w = draw.textlength(seg_text, font=cur_font)
            if is_hl and seg_text.strip():
                bg = [rx - pad_x, y + ref_bbox[1] - pad_y,
                      rx + seg_w + pad_x, y + ref_bbox[3] + pad_y]
                draw.rounded_rectangle(bg, radius=6, fill="#FFC107")
            rx += seg_w

        # ── 2nd pass: draw text ──
        rx = x_start
        for seg_text, is_hl in segments:
            color = "#111827" if is_hl else default_color
            draw.text((rx, y), seg_text, font=cur_font, fill=color)
            rx += draw.textlength(seg_text, font=cur_font)

        y += line_h + line_gap + 12

    return y


# ── Canvas helper functions ──────────────────────────────────────────

def split_rich_segments(text: str) -> list[tuple[str, bool]]:
    segments: list[tuple[str, bool]] = []
    for part in re.split(r"(\*\*.*?\*\*)", clean_text_for_font(text)):
        if not part:
            continue
        if part.startswith("**") and part.endswith("**"):
            segments.append((part[2:-2], True))
        else:
            segments.append((part, False))
    return segments or [(clean_text_for_font(text), False)]


def plain_from_segments(segments: list[tuple[str, bool]]) -> str:
    return "".join(text for text, _ in segments)


def fit_font_for_segments(draw, font_name: str, base_size: int, min_size: int, segments: list[tuple[str, bool]], max_width: int):
    use_shaped = shaped_text_available() and bool(font_path_for_name(font_name))
    size = base_size
    while size > min_size:
        font = load_font(font_name, size)
        width = (
            sum(shaped_text_width(font_name, size, text) for text, _ in segments)
            if use_shaped
            else sum(draw.textlength(text, font=font) for text, _ in segments)
        )
        if width <= max_width:
            return font, size, width
        size -= 2
    font = load_font(font_name, min_size)
    width = (
        sum(shaped_text_width(font_name, min_size, text) for text, _ in segments)
        if use_shaped
        else sum(draw.textlength(text, font=font) for text, _ in segments)
    )
    return font, min_size, width


def draw_centered_segment_line(
    draw,
    line: str,
    y: int,
    font_name: str,
    base_size: int,
    max_width: int,
    default_color: str,
    highlight_color: str,
    *,
    full_bg: str | None = None,
    highlight_bg: str | None = None,
    bg_pad_x: int = 18,
    bg_pad_y: int = 8,
    stroke_width: int = 2,
) -> int:
    segments = split_rich_segments(line)
    use_shaped = shaped_text_available() and bool(font_path_for_name(font_name))
    font, font_size, width = fit_font_for_segments(draw, font_name, base_size, 34, segments, max_width)
    plain_text = plain_from_segments(segments) or "Aygjป์"
    if use_shaped:
        _, shaped_bbox, _ = shaped_text_layout(font_name, font_size, plain_text)
        line_h = int(math.ceil(shaped_bbox[3] - shaped_bbox[1]))
    else:
        bbox = draw.textbbox((0, 0), plain_text, font=font, stroke_width=stroke_width)
        line_h = bbox[3] - bbox[1]
    x = int((CANVAS_SIZE[0] - width) / 2)

    if full_bg:
        if use_shaped:
            bg_box = (x - bg_pad_x, y - bg_pad_y, x + width + bg_pad_x, y + line_h + bg_pad_y)
        else:
            thai_top_pad = int(getattr(font, "size", base_size) * 0.08)
            bg_box = (
                x - bg_pad_x,
                y + bbox[1] - bg_pad_y - thai_top_pad,
                x + width + bg_pad_x,
                y + bbox[3] + bg_pad_y,
            )
        draw.rectangle(bg_box, fill=full_bg)

    cursor = x
    if highlight_bg:
        for text, is_highlight in segments:
            seg_w = shaped_text_width(font_name, font_size, text) if use_shaped else draw.textlength(text, font=font)
            if is_highlight and text.strip():
                if use_shaped:
                    bg_box = (cursor - 8, y - 4, cursor + seg_w + 8, y + line_h + 4)
                else:
                    thai_top_pad = int(getattr(font, "size", base_size) * 0.08)
                    bg_box = (
                        cursor - 8,
                        y + bbox[1] - 4 - thai_top_pad,
                        cursor + seg_w + 8,
                        y + bbox[3] + 4,
                    )
                draw.rectangle(bg_box, fill=highlight_bg)
            cursor += seg_w

    cursor = x
    for text, is_highlight in segments:
        fill = highlight_color if is_highlight else default_color
        if full_bg and is_highlight:
            fill = "#ffffff"
        if highlight_bg and is_highlight:
            fill = "#ffffff"
        if use_shaped:
            draw_shaped_text(draw._image, (int(round(cursor)), y), text, font_name, font_size, fill, stroke_width=stroke_width, stroke_fill="#000000")
            cursor += shaped_text_width(font_name, font_size, text)
        else:
            draw.text(
                (cursor, y),
                text,
                font=font,
                fill=fill,
                stroke_width=stroke_width,
                stroke_fill="#000000",
            )
            cursor += draw.textlength(text, font=font)

    return y + line_h


def normalize_headline_lines(item) -> list[str]:
    raw = str(getattr(item, "headline", "") or "")
    lines = [clean_text_for_font(line).strip() for line in raw.splitlines() if clean_text_for_font(line).strip()]
    if not lines:
        change = f"+{item.pct_change:.2f}%" if item.pct_change >= 0 else f"{item.pct_change:.2f}%"
        lines = [
            "หุ้นตัวนี้กำลังมาแรง",
            f"**{item.symbol}** วิ่ง {change} วันนี้",
            "โอกาสหรือความเสี่ยงที่ต้องจับตา",
        ]
    if len(lines) < 3:
        change = f"+{item.pct_change:.2f}%" if item.pct_change >= 0 else f"{item.pct_change:.2f}%"
        fallbacks = [
            f"**{item.symbol}** ขยับ {change}",
            "นักลงทุนต้องจับตารอบนี้",
            "อ่านก่อนตัดสินใจลงทุน",
        ]
        lines.extend(fallbacks[len(lines) - 1:])
    return lines[:3]


def headline_color_theme(item) -> dict[str, str]:
    theme_id = str(getattr(item, "headline_theme", "") or "classic").strip().lower()
    return HEADLINE_COLOR_THEMES.get(theme_id) or HEADLINE_COLOR_THEMES["classic"]


def paste_chart_cover(canvas, chart_path, area: tuple[int, int, int, int]) -> None:
    from PIL import Image

    x1, y1, x2, y2 = area
    target_w = x2 - x1
    target_h = y2 - y1
    chart = Image.open(chart_path).convert("RGB")
    scale = max(target_w / chart.width, target_h / chart.height)
    resized = chart.resize((int(chart.width * scale) + 1, int(chart.height * scale) + 1), Image.Resampling.LANCZOS)
    left = max(0, (resized.width - target_w) // 2)
    top = max(0, (resized.height - target_h) // 2)
    cropped = resized.crop((left, top, left + target_w, top + target_h))
    canvas.paste(cropped, (x1, y1))


def image_cover_from_bytes(image_bytes: bytes, size: tuple[int, int]):
    from PIL import Image

    img = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    target_w, target_h = size
    scale = max(target_w / img.width, target_h / img.height)
    resized = img.resize((int(img.width * scale) + 1, int(img.height * scale) + 1), Image.Resampling.LANCZOS)
    left = max(0, (resized.width - target_w) // 2)
    top = max(0, (resized.height - target_h) // 2)
    return resized.crop((left, top, left + target_w, top + target_h))


def set_canvas_size_for_ratio(image_ratio: str) -> None:
    global CANVAS_SIZE
    CANVAS_SIZE = SQUARE_CANVAS_SIZE if image_ratio == "square" else DEFAULT_CANVAS_SIZE


def is_square_canvas() -> bool:
    return CANVAS_SIZE[0] == CANVAS_SIZE[1]


def chart_overlay_seed(item) -> str:
    return (
        f"{getattr(item, 'symbol', '')}:"
        f"{getattr(item, 'company_name', '')}:"
        f"{getattr(item, 'pct_change', 0):.2f}:"
        f"{getattr(item, 'volume', 0)}:"
        f"{getattr(item, 'overlay_variant', '')}:"
        f"{str(getattr(item, 'headline', '') or '')[:80]}"
    )


def chart_overlay_positions(item, top_h: int) -> dict[str, tuple[int, int]]:
    """Pick matching news/meme slots so batch renders do not look cloned."""
    if is_square_canvas():
        slots = [
            {"news": (78, top_h - 210), "meme_image": (742, top_h - 226), "meme_sticker": (690, top_h - 168)},
            {"news": (454, top_h - 210), "meme_image": (76, top_h - 226), "meme_sticker": (66, top_h - 168)},
            {"news": (78, top_h - 330), "meme_image": (742, top_h - 220), "meme_sticker": (690, top_h - 168)},
            {"news": (454, top_h - 330), "meme_image": (76, top_h - 220), "meme_sticker": (66, top_h - 168)},
            {"news": (266, top_h - 210), "meme_image": (742, top_h - 390), "meme_sticker": (690, top_h - 332)},
            {"news": (78, top_h - 430), "meme_image": (742, top_h - 230), "meme_sticker": (690, top_h - 174)},
        ]
    else:
        slots = [
            {"news": (78, top_h - 228), "meme_image": (742, top_h - 262), "meme_sticker": (690, top_h - 182)},
            {"news": (454, top_h - 228), "meme_image": (76, top_h - 282), "meme_sticker": (66, top_h - 174)},
            {"news": (78, top_h - 350), "meme_image": (742, top_h - 238), "meme_sticker": (690, top_h - 174)},
            {"news": (454, top_h - 350), "meme_image": (76, top_h - 238), "meme_sticker": (66, top_h - 174)},
            {"news": (266, top_h - 228), "meme_image": (742, top_h - 390), "meme_sticker": (690, top_h - 332)},
            {"news": (78, top_h - 430), "meme_image": (742, top_h - 250), "meme_sticker": (690, top_h - 184)},
        ]
    try:
        variant = int(getattr(item, "overlay_variant", 0) or 0)
    except (TypeError, ValueError):
        variant = 0
    if variant > 0:
        return slots[(variant - 1) % len(slots)]
    return slots[stable_index(chart_overlay_seed(item), len(slots))]


def chart_top_h_for_item(item) -> int:
    return int(getattr(item, "chart_top_h", 650 if is_square_canvas() else 810))


def text_lines_to_width(draw, text: str, font, max_width: int, max_lines: int) -> list[str]:
    words = text.split()
    lines: list[str] = []
    current = ""
    for word in words:
        if draw.textlength(word, font=font) > max_width:
            if current:
                lines.append(current)
                current = ""
            for piece in wrap_long_token(draw, word, font, max_width):
                if len(lines) >= max_lines:
                    break
                lines.append(piece)
            if len(lines) >= max_lines:
                break
            continue
        candidate = f"{current} {word}".strip()
        if draw.textlength(candidate, font=font) <= max_width:
            current = candidate
            continue
        if current:
            lines.append(current)
        current = word
        if len(lines) >= max_lines:
            break
    if current and len(lines) < max_lines:
        lines.append(current)
    if len(lines) == max_lines and words:
        while lines[-1] and draw.textlength(lines[-1] + "...", font=font) > max_width:
            lines[-1] = lines[-1][:-1]
        if not lines[-1].endswith("..."):
            lines[-1] = lines[-1].rstrip() + "..."
    return lines


def draw_news_card(draw, canvas, item, top_h: int) -> None:
    from PIL import Image, ImageDraw

    news = next((entry for entry in item.news_items if entry.image_bytes), None)
    news = news or (item.news_items[0] if item.news_items else None)
    if not news:
        return

    card_x, card_y = chart_overlay_positions(item, top_h)["news"]
    card_w, card_h = 548, 164
    img_w, img_h = 174, 116
    overlay = Image.new("RGBA", CANVAS_SIZE, (0, 0, 0, 0))
    odraw = ImageDraw.Draw(overlay)
    shadow = (card_x + 8, card_y + 10, card_x + card_w + 8, card_y + card_h + 10)
    odraw.rounded_rectangle(shadow, radius=10, fill=(0, 0, 0, 130))
    odraw.rounded_rectangle((card_x, card_y, card_x + card_w, card_y + card_h), radius=10, fill=(4, 8, 13, 232), outline=(255, 255, 255, 210), width=3)
    canvas_rgba = canvas.convert("RGBA")
    canvas_rgba.alpha_composite(overlay)
    canvas.paste(canvas_rgba.convert("RGB"))

    image_x, image_y = card_x + 18, card_y + 24
    if news.image_bytes:
        try:
            news_img = image_cover_from_bytes(news.image_bytes, (img_w, img_h))
            mask = Image.new("L", (img_w, img_h), 0)
            mask_draw = ImageDraw.Draw(mask)
            mask_draw.rounded_rectangle((0, 0, img_w, img_h), radius=7, fill=255)
            canvas.paste(news_img, (image_x, image_y), mask)
        except Exception:
            draw.rounded_rectangle((image_x, image_y, image_x + img_w, image_y + img_h), radius=7, fill="#111827", outline="#2f3a45", width=2)
    else:
        draw.rounded_rectangle((image_x, image_y, image_x + img_w, image_y + img_h), radius=7, fill="#111827", outline="#2f3a45", width=2)
        news_font = load_font("Kanit-Bold.ttf", 38)
        label = "NEWS"
        label_w = draw.textlength(label, font=news_font)
        draw.text((image_x + (img_w - label_w) / 2, image_y + 34), label, font=news_font, fill="#ffffff")

    text_x = image_x + img_w + 20
    label_font = load_font("Sarabun-Bold.ttf", 27)
    detail_font = load_font("Sarabun-Bold.ttf", 26)
    source_font = load_font("Sarabun-Bold.ttf", 17)
    label, detail = news_overlay_copy(item, news)
    draw.text((text_x, card_y + 22), label, font=label_font, fill="#fff200", stroke_width=1, stroke_fill="#000000")
    for idx, line in enumerate(text_lines_to_width(draw, detail, detail_font, 300, 2)):
        draw.text((text_x, card_y + 61 + idx * 34), line, font=detail_font, fill="#ffffff", stroke_width=1, stroke_fill="#000000")
    source = f"ที่มา: {news.source or 'ข่าวล่าสุด'}"
    draw.text((text_x, card_y + 128), source, font=source_font, fill="#cbd5e1")


def clean_meme_line(text: str, max_chars: int) -> str:
    text = re.sub(r"^[\-\d\.\)\s]+", "", str(text or "")).strip()
    text = text.replace('"', "").replace("'", "").strip()
    text = re.sub(r"\s+", " ", text)
    if len(text) > max_chars:
        text = text[:max_chars].rstrip()
    return text


def fallback_meme_copy(item: StockItem) -> tuple[str, str]:
    business = company_share_angle(item)
    negative = [
        ("ใครกดขายก่อน?", f"{item.symbol} ยังมีเรื่องให้ลุ้น"),
        ("แดงแต่คนดูเยอะ", "วอลุ่มแบบนี้ต้องตาม"),
        ("เด้งหรือดิ่งต่อ?", "อย่าเพิ่งรีบสรุป"),
        ("ตลาดกำลังลังเล", f"{business} ยังเป็นประเด็น"),
        ("รอจังหวะก่อน", "ข่าวเริ่มเขย่าราคา"),
    ]
    positive = [
        ("เขียวแบบมีสตอรี่", "รอบนี้ไม่ใช่เด้งเฉยๆ"),
        ("เข้าไม่ทันแล้ว?", "แรงซื้อมาแบบมีเหตุผล"),
        ("ตลาดเริ่มหันมอง", f"{business} กำลังร้อน"),
        ("ถืออยู่ยิ้มไหม", "โมเมนตัมยังมา"),
        ("รอบนี้มีเรื่อง", "ข่าวหนุนราคาเริ่มชัด"),
    ]
    neutral = [
        ("นิ่งแต่ไม่น่าเบื่อ", "ตลาดรอสัญญาณใหม่"),
        ("เงียบแบบมีนัย", "ข่าวรอบนี้ต้องอ่าน"),
        ("ยังไม่จบง่ายๆ", "วอลุ่มกำลังบอกอะไร"),
    ]
    pool = negative if item.pct_change < 0 else positive if item.pct_change > 0 else neutral
    return pool[stable_index(f"{item.symbol}:{item.company_name}:{item.pct_change}", len(pool))]


def meme_copy(item: StockItem) -> tuple[str, str]:
    title = clean_meme_line(getattr(item, "meme_title", ""), 18)
    subtitle = clean_meme_line(getattr(item, "meme_subtitle", ""), 24)
    if title and subtitle:
        return title, subtitle
    return fallback_meme_copy(item)


def generate_meme_copy(item: StockItem, creds: AppCreds, model: str) -> tuple[str, str]:
    fallback = fallback_meme_copy(item)
    if not creds.openrouter_key:
        return fallback
    news_title = item.news_items[0].title if item.news_items else (item.news[0] if item.news else "")
    prompt = (
        "คุณคือครีเอเตอร์มีมการเงินไทย เขียนข้อความสั้นมากสำหรับ sticker บนภาพหุ้น\n"
        "ต้องอ่านแล้วแชร์ได้ ไม่ซ้ำ ไม่ generic ไม่ใช้คำว่า ใจเย็นก่อน\n"
        "ตอบแค่ 2 บรรทัดเท่านั้น ห้ามใส่เลขข้อ ห้ามใส่คำอธิบาย\n"
        "บรรทัด 1: punchline ไทย 8-18 ตัวอักษร\n"
        "บรรทัด 2: คำต่อท้ายไทย 10-24 ตัวอักษร\n\n"
        f"หุ้น: {item.symbol}\nบริษัท: {item.company_name}\n"
        f"ราคาเปลี่ยน: {item.pct_change:+.2f}%\nVolume: {item.volume:,}\n"
        f"ธุรกิจ: {company_share_angle(item)}\nพาดหัวภาพ: {item.headline}\nข่าว: {news_title}\n"
    )
    try:
        raw = chat_completion(
            api_key=creds.openrouter_key,
            base_url="https://openrouter.ai/api/v1/chat/completions",
            model=model,
            prompt=prompt,
            headers={"HTTP-Referer": "http://localhost", "X-Title": "BulkVideoCreatorApp"},
        )
        lines = [clean_meme_line(line, 28) for line in raw.splitlines() if clean_meme_line(line, 28)]
        if len(lines) >= 2:
            return clean_meme_line(lines[0], 18), clean_meme_line(lines[1], 24)
    except Exception as exc:
        print(f"[WARN] AI meme copy failed for {item.symbol}: {exc}", file=sys.stderr)
    return fallback


def meme_search_query(item: StockItem) -> str:
    if item.pct_change <= -4:
        return "stock market crash reaction"
    if item.pct_change < 0:
        return "confused reaction meme"
    if item.pct_change >= 6:
        return "stonks money reaction"
    if item.pct_change > 0:
        return "excited reaction meme"
    return "thinking reaction meme"


def fetch_giphy_meme(item: StockItem, api_key: str) -> bytes | None:
    if not api_key:
        print(f"[WARN] GIPHY API key missing for {item.symbol}; using local sticker.", file=sys.stderr)
        return None
    session = requests_session()
    try:
        res = session.get(
            "https://api.giphy.com/v1/gifs/search",
            params={
                "api_key": api_key,
                "q": meme_search_query(item),
                "limit": 20,
                "rating": "pg-13",
                "lang": "en",
            },
            timeout=14,
        )
        if res.status_code != 200:
            print(f"[WARN] GIPHY search failed for {item.symbol}: HTTP {res.status_code} {res.text[:160]}", file=sys.stderr)
            return None
        gifs = res.json().get("data") or []
        if not isinstance(gifs, list) or not gifs:
            print(f"[WARN] GIPHY returned no meme results for {item.symbol}; query={meme_search_query(item)!r}", file=sys.stderr)
            return None
        index = stable_index(item.symbol, len(gifs))
        ordered = gifs[index:] + gifs[:index]
        last_status = ""
        for gif in ordered:
            images = gif.get("images") if isinstance(gif, dict) else {}
            candidates = [
                images.get("original_still", {}).get("url") if isinstance(images, dict) else "",
                images.get("fixed_height_still", {}).get("url") if isinstance(images, dict) else "",
                images.get("fixed_width_still", {}).get("url") if isinstance(images, dict) else "",
                images.get("downsized_still", {}).get("url") if isinstance(images, dict) else "",
                images.get("480w_still", {}).get("url") if isinstance(images, dict) else "",
            ]
            for url in [candidate for candidate in candidates if candidate]:
                img_res = session.get(url, timeout=14)
                if img_res.status_code == 200 and is_supported_image_bytes(img_res.content):
                    print(f"[INFO] GIPHY meme selected for {item.symbol}")
                    return img_res.content
                last_status = f"HTTP {img_res.status_code} {url[:90]}"
        if last_status:
            print(f"[WARN] GIPHY image downloads failed for {item.symbol}: {last_status}", file=sys.stderr)
    except Exception as exc:
        print(f"[WARN] GIPHY meme failed for {item.symbol}: {exc}", file=sys.stderr)
    return None


def local_meme_bytes(path: str) -> bytes | None:
    clean = str(path or "").strip()
    if not clean:
        return None
    try:
        image_path = Path(clean).expanduser()
        if image_path.exists() and image_path.is_dir():
            candidates = sorted(
                p for p in image_path.iterdir()
                if p.is_file() and p.suffix.lower() in {".png", ".jpg", ".jpeg", ".webp", ".gif"}
            )
            if candidates:
                seed = str(getattr(local_meme_bytes, "_seed_symbol", "")) or clean
                image_path = candidates[stable_index(seed, len(candidates))]
        if image_path.exists() and image_path.is_file():
            data = image_path.read_bytes()
            if is_supported_image_bytes(data):
                return data
    except Exception as exc:
        print(f"[WARN] Local meme image failed: {clean} ({exc})", file=sys.stderr)
    return None


def draw_meme_image_card(draw, canvas, item, image_bytes: bytes, source_label: str = "") -> None:
    from PIL import Image, ImageDraw

    top_h = chart_top_h_for_item(item)
    x, y = chart_overlay_positions(item, top_h)["meme_image"]
    if is_square_canvas():
        w, h = 286, 190
    else:
        w, h = 286, 198
    overlay = Image.new("RGBA", CANVAS_SIZE, (0, 0, 0, 0))
    odraw = ImageDraw.Draw(overlay)
    odraw.rounded_rectangle((x + 8, y + 9, x + w + 8, y + h + 9), radius=16, fill=(0, 0, 0, 150))
    canvas_rgba = canvas.convert("RGBA")
    canvas_rgba.alpha_composite(overlay)
    canvas.paste(canvas_rgba.convert("RGB"))

    try:
        meme_img = image_cover_from_bytes(image_bytes, (w, h))
        mask = Image.new("L", (w, h), 0)
        mask_draw = ImageDraw.Draw(mask)
        mask_draw.rounded_rectangle((0, 0, w, h), radius=14, fill=255)
        canvas.paste(meme_img, (x, y), mask)
    except Exception:
        return


def draw_meme_sticker(draw, canvas, item) -> None:
    from PIL import Image, ImageDraw

    top_h = chart_top_h_for_item(item)
    x, y = chart_overlay_positions(item, top_h)["meme_sticker"]
    w, h = 352, 126
    variant = stable_index(f"{item.symbol}:{item.company_name}", 6)
    palettes = [
        ((255, 255, 255, 246), (255, 242, 0, 255), "#ffb84d"),
        ((250, 245, 255, 246), (168, 85, 247, 255), "#facc15"),
        ((236, 253, 245, 246), (34, 197, 94, 255), "#fde68a"),
        ((239, 246, 255, 246), (59, 130, 246, 255), "#fed7aa"),
        ((255, 247, 237, 246), (249, 115, 22, 255), "#fcd34d"),
        ((254, 242, 242, 246), (239, 68, 68, 255), "#fdba74"),
    ]
    fill, outline, face_fill = palettes[variant]
    overlay = Image.new("RGBA", CANVAS_SIZE, (0, 0, 0, 0))
    odraw = ImageDraw.Draw(overlay)
    odraw.rounded_rectangle((x + 10, y + 11, x + w + 10, y + h + 11), radius=20, fill=(0, 0, 0, 170))
    odraw.rounded_rectangle((x, y, x + w, y + h), radius=20, fill=fill, outline=outline, width=5)
    odraw.polygon(((x + 42, y + h), (x + 84, y + h), (x + 62, y + h + 20)), fill=outline)
    canvas_rgba = canvas.convert("RGBA")
    canvas_rgba.alpha_composite(overlay)
    canvas.paste(canvas_rgba.convert("RGB"))

    face_cx, face_cy, r = x + 63, y + 63, 42
    draw.ellipse((face_cx - r, face_cy - r, face_cx + r, face_cy + r), fill=face_fill, outline="#111827", width=3)
    eye_y = face_cy - 8
    expression = variant if item.pct_change >= 0 else (variant + 3) % 6
    if expression in {0, 3}:
        draw.line((face_cx - 18, eye_y - 4, face_cx - 8, eye_y + 6), fill="#111827", width=4)
        draw.line((face_cx + 8, eye_y + 6, face_cx + 18, eye_y - 4), fill="#111827", width=4)
        draw.arc((face_cx - 17, face_cy + 12, face_cx + 17, face_cy + 38), 200, 340, fill="#111827", width=4)
    elif expression in {1, 4}:
        draw.ellipse((face_cx - 20, eye_y - 9, face_cx - 6, eye_y + 9), fill="#111827")
        draw.ellipse((face_cx + 6, eye_y - 9, face_cx + 20, eye_y + 9), fill="#111827")
        draw.ellipse((face_cx - 10, face_cy + 14, face_cx + 10, face_cy + 34), outline="#111827", width=4)
        if item.pct_change < 0:
            draw.arc((face_cx + 24, face_cy - 16, face_cx + 44, face_cy + 18), 80, 260, fill="#38bdf8", width=4)
    else:
        draw.ellipse((face_cx - 18, eye_y - 4, face_cx - 8, eye_y + 6), fill="#111827")
        draw.ellipse((face_cx + 8, eye_y - 4, face_cx + 18, eye_y + 6), fill="#111827")
        draw.arc((face_cx - 18, face_cy, face_cx + 18, face_cy + 28), 10, 170, fill="#111827", width=4)

    title, subtitle = meme_copy(item)
    title_font = fit_font_to_width(draw, title, "Sarabun-Bold.ttf", 36, 24, 220)
    sub_font = fit_font_to_width(draw, subtitle, "Sarabun-Bold.ttf", 27, 20, 220)
    tx = x + 126
    draw.text((tx, y + 26), title, font=title_font, fill="#111827")
    draw.text((tx, y + 72), subtitle, font=sub_font, fill="#334155")


def draw_selected_meme(draw, canvas, item) -> None:
    source = str(getattr(item, "meme_source", "local") or "local").strip().lower()
    if source == "giphy":
        image_bytes = fetch_giphy_meme(item, str(getattr(item, "giphy_api_key", "") or ""))
        if image_bytes:
            draw_meme_image_card(draw, canvas, item, image_bytes, "Powered by GIPHY")
            return
    setattr(local_meme_bytes, "_seed_symbol", f"{item.symbol}:{item.company_name}:{item.pct_change}")
    image_bytes = local_meme_bytes(str(getattr(item, "local_meme_path", "") or ""))
    if image_bytes:
        draw_meme_image_card(draw, canvas, item, image_bytes, "Local meme")
        return
    draw_meme_sticker(draw, canvas, item)


def draw_page_credit(draw, item) -> None:
    credit = str(getattr(item, "page_credit", "") or "").strip()
    if not credit:
        return
    if len(credit) > 72:
        credit = credit[:69].rstrip() + "..."
    text = credit if credit.startswith("@") or credit.startswith("เครดิต:") else f"เครดิต: {credit}"
    font = fit_font_to_width(draw, text, "Sarabun-Bold.ttf", 28, 18, 920)
    text_w = draw.textlength(text, font=font)
    x = int((CANVAS_SIZE[0] - text_w) / 2)
    y = CANVAS_SIZE[1] - 58
    draw.text((x, y), text, font=font, fill="#cbd5e1", stroke_width=1, stroke_fill="#000000")


def draw_chart_overlay(draw, canvas, item, logo_bytes, top_h: int) -> None:
    canvas_w = CANVAS_SIZE[0]
    logo = image_from_logo_bytes(logo_bytes, item.symbol)
    box_x1, box_y1, box_h = 34, 34, 126
    logo_size = 84
    text_x = box_x1 + 120
    symbol_font = load_font("Kanit-Bold.ttf", 42)
    company_font = fit_font_to_width(draw, item.company_name, "Sarabun-Bold.ttf", 22, 16, 420)
    company_text = ellipsize_to_width(draw, item.company_name, company_font, 420)
    text_w = max(draw.textlength(item.symbol, font=symbol_font), draw.textlength(company_text, font=company_font))
    box_w = int(min(610, max(330, 146 + text_w)))
    draw.rounded_rectangle((box_x1, box_y1, box_x1 + box_w, box_y1 + box_h), radius=8, fill="#050505", outline="#2f3a45", width=2)
    logo_resized = logo.resize((logo_size, logo_size))
    canvas.paste(logo_resized, (box_x1 + 20, box_y1 + 21), logo_resized)
    draw.text((text_x, box_y1 + 29), item.symbol, font=symbol_font, fill="#ffffff")
    draw.text((text_x + 2, box_y1 + 78), company_text, font=company_font, fill="#d7dee8")

    accent = BULL if item.pct_change >= 0 else BEAR
    metric = f"+{item.pct_change:.2f}%" if item.pct_change >= 0 else f"{item.pct_change:.2f}%"
    metric_font = load_font("Kanit-Bold.ttf", 46)
    price_font = load_font("Prompt-Bold.ttf", 24)
    metric_w = draw.textlength(metric, font=metric_font)
    price = f"${item.price:,.2f}"
    price_w = draw.textlength(price, font=price_font)
    box_w = int(max(metric_w, price_w) + 58)
    right = canvas_w - 34
    draw.rounded_rectangle((canvas_w - box_w - 34, 34, right, 150), radius=8, fill="#050505", outline=accent, width=3)
    draw.text((canvas_w - box_w - 5, 46), metric, font=metric_font, fill=accent)
    draw.text((right - price_w - 24, 104), price, font=price_font, fill="#ffffff")

    for i in range(90):
        alpha = i / 89
        shade = int(13 * (1 - alpha))
        y = top_h - 90 + i
        draw.line((0, y, canvas_w, y), fill=(shade, shade, shade))
    draw_news_card(draw, canvas, item, top_h)


def compose_canvas_viral(item, chart_path, logo_bytes, output_path):
    from PIL import Image, ImageDraw

    canvas_w, canvas_h = CANVAS_SIZE
    top_h = 810
    setattr(item, "chart_top_h", top_h)
    canvas = Image.new("RGB", CANVAS_SIZE, "#000000")
    draw = ImageDraw.Draw(canvas)
    paste_chart_cover(canvas, chart_path, (0, 0, canvas_w, top_h))
    draw_chart_overlay(draw, canvas, item, logo_bytes, top_h)

    draw.rectangle((0, top_h, canvas_w, canvas_h), fill="#000000")
    accent = BULL if item.pct_change >= 0 else BEAR
    draw.rectangle((0, top_h, canvas_w, top_h + 8), fill=accent)

    lines = normalize_headline_lines(item)
    theme = headline_color_theme(item)
    draw_centered_segment_line(
        draw,
        lines[0],
        top_h + 58,
        VIRAL_HEADLINE_FONT,
        58,
        940,
        theme["line1_text"],
        theme["line1_text"],
        full_bg=theme["line1_bg"],
        bg_pad_x=24,
        bg_pad_y=12,
        stroke_width=0,
    )
    draw_centered_segment_line(
        draw,
        lines[1],
        top_h + 174,
        VIRAL_HEADLINE_FONT,
        64,
        1010,
        theme["line2_text"],
        theme["line2_highlight"],
        stroke_width=2,
    )
    draw_centered_segment_line(
        draw,
        lines[2],
        top_h + 292,
        VIRAL_HEADLINE_FONT,
        58,
        1010,
        theme["line3_text"],
        theme["line3_text"],
        full_bg=theme["line3_bg"],
        bg_pad_x=20,
        bg_pad_y=12,
        stroke_width=0,
    )
    if getattr(item, "meme_overlay", False):
        draw_selected_meme(draw, canvas, item)
    draw_page_credit(draw, item)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_path, "PNG", optimize=True)
    return output_path


def compose_canvas_viral_square(item, chart_path, logo_bytes, output_path):
    from PIL import Image, ImageDraw

    canvas_w, canvas_h = CANVAS_SIZE
    top_h = 650
    setattr(item, "chart_top_h", top_h)
    canvas = Image.new("RGB", CANVAS_SIZE, "#000000")
    draw = ImageDraw.Draw(canvas)
    paste_chart_cover(canvas, chart_path, (0, 0, canvas_w, top_h))
    draw_chart_overlay(draw, canvas, item, logo_bytes, top_h)

    draw.rectangle((0, top_h, canvas_w, canvas_h), fill="#000000")
    accent = BULL if item.pct_change >= 0 else BEAR
    draw.rectangle((0, top_h, canvas_w, top_h + 8), fill=accent)

    lines = normalize_headline_lines(item)
    theme = headline_color_theme(item)
    draw_centered_segment_line(
        draw,
        lines[0],
        top_h + 48,
        VIRAL_HEADLINE_FONT,
        52,
        900,
        theme["line1_text"],
        theme["line1_text"],
        full_bg=theme["line1_bg"],
        bg_pad_x=22,
        bg_pad_y=10,
        stroke_width=0,
    )
    draw_centered_segment_line(
        draw,
        lines[1],
        top_h + 142,
        VIRAL_HEADLINE_FONT,
        56,
        1010,
        theme["line2_text"],
        theme["line2_highlight"],
        stroke_width=2,
    )
    draw_centered_segment_line(
        draw,
        lines[2],
        top_h + 244,
        VIRAL_HEADLINE_FONT,
        50,
        1010,
        theme["line3_text"],
        theme["line3_text"],
        full_bg=theme["line3_bg"],
        bg_pad_x=18,
        bg_pad_y=10,
        stroke_width=0,
    )
    if getattr(item, "meme_overlay", False):
        draw_selected_meme(draw, canvas, item)
    draw_page_credit(draw, item)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_path, "PNG", optimize=True)
    return output_path

def _draw_header(draw, canvas, item, logo_bytes, title_font, company_font, metric_font, price_font, bg_fill, outline_color, text_color, muted_color):
    draw.rounded_rectangle((44, 44, 1036, 216), radius=8, fill=bg_fill, outline=outline_color, width=2)
    logo = image_from_logo_bytes(logo_bytes, item.symbol)
    canvas.paste(logo, (74, 70), logo)
    draw.text((218, 68), item.symbol, font=title_font, fill=text_color)
    company = item.company_name[:34] + ("..." if len(item.company_name) > 34 else "")
    draw.text((222, 145), company, font=company_font, fill=muted_color)
    metric = f"+{item.pct_change:.2f}%" if item.pct_change >= 0 else f"{item.pct_change:.2f}%"
    m_bbox = draw.textbbox((0, 0), metric, font=metric_font)
    draw.text((1002 - (m_bbox[2] - m_bbox[0]), 68), metric, font=metric_font, fill=BULL if item.pct_change >= 0 else BEAR)
    price = f"${item.price:,.2f}"
    p_bbox = draw.textbbox((0, 0), price, font=price_font)
    draw.text((1002 - (p_bbox[2] - p_bbox[0]), 145), price, font=price_font, fill=text_color)


def _paste_chart(draw, canvas, chart_path, chart_area, bg_fill, outline_color):
    from PIL import Image
    x1, y1, x2, y2 = chart_area
    draw.rounded_rectangle(chart_area, radius=8, fill=bg_fill, outline=outline_color, width=2)
    chart = Image.open(chart_path).convert("RGB")
    chart.thumbnail((x2 - x1 - 12, y2 - y1 - 12), Image.Resampling.LANCZOS)
    canvas.paste(chart, (x1 + (x2 - x1 - chart.width) // 2, y1 + (y2 - y1 - chart.height) // 2))


# ── Canvas Style 1: Classic Dark ─────────────────────────────────────
def compose_canvas_classic(item, chart_path, logo_bytes, output_path):
    from PIL import Image, ImageDraw
    canvas = Image.new("RGB", CANVAS_SIZE, BG)
    draw = ImageDraw.Draw(canvas)
    draw.rectangle((0, 0, 1080, 10), fill=BULL)
    _draw_header(draw, canvas, item, logo_bytes,
                 load_font("Kanit-Bold.ttf", 66), load_font("Prompt-Regular.ttf", 30),
                 load_font("Kanit-Bold.ttf", 60), load_font("Prompt-Bold.ttf", 34),
                 "#161b22", "#30363d", TEXT, MUTED)
    _paste_chart(draw, canvas, chart_path, (44, 236, 1036, 948), "#0b0f14", "#30363d")
    hl = getattr(item, 'headline', f"อัปเดตหุ้น **{item.symbol}**")
    draw_rich_text_wrapped(draw, hl, (74, 1000), load_font("Kanit-Bold.ttf", 60), TEXT, "#FFC107", 932, line_gap=16)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_path, "PNG", optimize=True)
    return output_path


# ── Canvas Style 2: Neon Glow ────────────────────────────────────────
def compose_canvas_neon(item, chart_path, logo_bytes, output_path):
    from PIL import Image, ImageDraw
    canvas = Image.new("RGB", CANVAS_SIZE, "#0a0a1a")
    draw = ImageDraw.Draw(canvas)
    for yp in range(1350):
        draw.line([(0, yp), (1080, yp)], fill=(int(10 + yp/90), int(10 + yp/270), int(26 + yp/67)))
    accent = "#00d4ff" if item.pct_change >= 0 else "#ff4d6a"
    draw.rectangle((0, 0, 1080, 6), fill=accent)
    draw.rectangle((0, 1344, 1080, 1350), fill=accent)
    _draw_header(draw, canvas, item, logo_bytes,
                 load_font("Kanit-Bold.ttf", 66), load_font("Prompt-Regular.ttf", 30),
                 load_font("Kanit-Bold.ttf", 60), load_font("Prompt-Bold.ttf", 34),
                 "#12122a", "#2a2a5a", "#e0e0ff", "#7a7aaa")
    _paste_chart(draw, canvas, chart_path, (44, 236, 1036, 948), "#08081a", "#1a1a4a")
    hl = getattr(item, 'headline', f"อัปเดตหุ้น **{item.symbol}**")
    draw_rich_text_wrapped(draw, hl, (74, 1000), load_font("Kanit-Bold.ttf", 58), "#e0e0ff", "#FFC107", 932, line_gap=16)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_path, "PNG", optimize=True)
    return output_path


# ── Canvas Style 3: Clean White ──────────────────────────────────────
def compose_canvas_clean(item, chart_path, logo_bytes, output_path):
    from PIL import Image, ImageDraw
    canvas = Image.new("RGB", CANVAS_SIZE, "#f5f5f5")
    draw = ImageDraw.Draw(canvas)
    accent = BULL if item.pct_change >= 0 else BEAR
    draw.rectangle((0, 0, 1080, 8), fill=accent)
    tf = load_font("Kanit-Bold.ttf", 66)
    cf = load_font("Prompt-Regular.ttf", 30)
    mf = load_font("Kanit-Bold.ttf", 60)
    pf = load_font("Prompt-Bold.ttf", 34)
    draw.rounded_rectangle((44, 44, 1036, 216), radius=12, fill="#ffffff", outline="#e0e0e0", width=2)
    logo = image_from_logo_bytes(logo_bytes, item.symbol)
    canvas.paste(logo, (74, 70), logo)
    draw.text((218, 68), item.symbol, font=tf, fill="#1a1a2e")
    co = item.company_name[:34] + ("..." if len(item.company_name) > 34 else "")
    draw.text((222, 145), co, font=cf, fill="#666680")
    met = f"+{item.pct_change:.2f}%" if item.pct_change >= 0 else f"{item.pct_change:.2f}%"
    mb = draw.textbbox((0,0), met, font=mf)
    draw.text((1002-(mb[2]-mb[0]), 68), met, font=mf, fill=accent)
    pr = f"${item.price:,.2f}"
    pb = draw.textbbox((0,0), pr, font=pf)
    draw.text((1002-(pb[2]-pb[0]), 145), pr, font=pf, fill="#1a1a2e")
    _paste_chart(draw, canvas, chart_path, (44, 236, 1036, 948), "#ffffff", "#e0e0e0")
    hl = getattr(item, 'headline', f"อัปเดตหุ้น **{item.symbol}**")
    draw_rich_text_wrapped(draw, hl, (74, 1000), load_font("Kanit-Bold.ttf", 56), "#1a1a2e", "#FFC107", 932, line_gap=16)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_path, "PNG", optimize=True)
    return output_path


# ── Canvas Style 4: Bold Impact ──────────────────────────────────────
def compose_canvas_bold(item, chart_path, logo_bytes, output_path):
    from PIL import Image, ImageDraw
    canvas = Image.new("RGB", CANVAS_SIZE, "#111111")
    draw = ImageDraw.Draw(canvas)
    accent = BULL if item.pct_change >= 0 else BEAR
    draw.rectangle((0, 0, 12, 1350), fill=accent)
    draw.rectangle((1068, 0, 1080, 1350), fill=accent)
    tf = load_font("Kanit-Bold.ttf", 72)
    cf = load_font("Prompt-Regular.ttf", 28)
    mf = load_font("Kanit-Bold.ttf", 80)
    pf = load_font("Prompt-Bold.ttf", 36)
    met = f"+{item.pct_change:.2f}%" if item.pct_change >= 0 else f"{item.pct_change:.2f}%"
    mb = draw.textbbox((0,0), met, font=mf)
    draw.text(((1080-(mb[2]-mb[0]))//2, 50), met, font=mf, fill=accent)
    sb = draw.textbbox((0,0), item.symbol, font=tf)
    draw.text(((1080-(sb[2]-sb[0]))//2, 155), item.symbol, font=tf, fill="#ffffff")
    co = item.company_name[:40] + ("..." if len(item.company_name) > 40 else "")
    cb = draw.textbbox((0,0), co, font=cf)
    draw.text(((1080-(cb[2]-cb[0]))//2, 240), co, font=cf, fill="#888888")
    pr = f"${item.price:,.2f}"
    pb = draw.textbbox((0,0), pr, font=pf)
    draw.text(((1080-(pb[2]-pb[0]))//2, 285), pr, font=pf, fill="#cccccc")
    _paste_chart(draw, canvas, chart_path, (44, 350, 1036, 880), "#0a0a0a", "#333333")
    hl = getattr(item, 'headline', f"อัปเดตหุ้น **{item.symbol}**")
    draw_rich_text_wrapped(draw, hl, (74, 930), load_font("Kanit-Bold.ttf", 64), "#ffffff", "#FFC107", 932, line_gap=16)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_path, "PNG", optimize=True)
    return output_path


CANVAS_STYLES = {
    "viral": compose_canvas_viral,
    "classic": compose_canvas_classic,
    "neon": compose_canvas_neon,
    "clean": compose_canvas_clean,
    "bold": compose_canvas_bold,
}


def compose_canvas(
    item: StockItem,
    chart_path: Path,
    logo_bytes: bytes | None,
    output_path: Path,
    style: str = "classic",
    image_ratio: str = "default",
    headline_theme: str = "classic",
    meme_overlay: bool = False,
    page_credit: str = "",
    meme_source: str = "local",
    giphy_api_key: str = "",
    local_meme_path: str = "",
) -> Path:
    set_canvas_size_for_ratio(image_ratio)
    setattr(item, "headline_theme", headline_theme)
    setattr(item, "meme_overlay", meme_overlay)
    setattr(item, "page_credit", page_credit)
    setattr(item, "meme_source", meme_source)
    setattr(item, "giphy_api_key", giphy_api_key)
    setattr(item, "local_meme_path", local_meme_path)
    if style == "viral" and image_ratio == "square":
        fn = compose_canvas_viral_square
    else:
        fn = CANVAS_STYLES.get(style, compose_canvas_classic)
    return fn(item, chart_path, logo_bytes, output_path)




    from PIL import Image, ImageDraw
    canvas = Image.new("RGB", CANVAS_SIZE, BG)
    draw = ImageDraw.Draw(canvas)
    title_font = load_font("Kanit-Bold.ttf", 66)
    company_font = load_font("Prompt-Regular.ttf", 30)
    metric_font = load_font("Kanit-Bold.ttf", 60)
    price_font = load_font("Prompt-Bold.ttf", 34)
    headline_font = load_font("Kanit-Bold.ttf", 60)

    draw.rectangle((0, 0, 1080, 10), fill=BULL)
    _draw_header(draw, canvas, item, logo_bytes, title_font, company_font, metric_font, price_font,
                 "#161b22", "#30363d", TEXT, MUTED)
    _paste_chart(draw, canvas, chart_path, (44, 236, 1036, 948), "#0b0f14", "#30363d")

    headline_text = getattr(item, 'headline', f"อัปเดตหุ้น **{item.symbol}** วันนี้")
    draw_rich_text_wrapped(draw, headline_text, (74, 1000), headline_font, TEXT, "#FFC107", 932, line_gap=16)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_path, "PNG", optimize=True)
    return output_path


def compose_canvas_neon(item, chart_path, logo_bytes, output_path):
    """Style 2: Neon glow — deep purple/blue gradient background with cyan/magenta accents."""
    from PIL import Image, ImageDraw
    canvas = Image.new("RGB", CANVAS_SIZE, "#0a0a1a")
    draw = ImageDraw.Draw(canvas)

    # Gradient background effect (top to bottom)
    for y_pos in range(1350):
        r = int(10 + (y_pos / 1350) * 15)
        g = int(10 + (y_pos / 1350) * 5)
        b = int(26 + (y_pos / 1350) * 20)
        draw.line([(0, y_pos), (1080, y_pos)], fill=(r, g, b))

    # Accent bars
    accent = "#00d4ff" if item.pct_change >= 0 else "#ff4d6a"
    draw.rectangle((0, 0, 1080, 6), fill=accent)
    draw.rectangle((0, 1344, 1080, 1350), fill=accent)

    title_font = load_font("Kanit-Bold.ttf", 66)
    company_font = load_font("Prompt-Regular.ttf", 30)
    metric_font = load_font("Kanit-Bold.ttf", 60)
    price_font = load_font("Prompt-Bold.ttf", 34)
    headline_font = load_font("Kanit-Bold.ttf", 58)

    _draw_header(draw, canvas, item, logo_bytes, title_font, company_font, metric_font, price_font,
                 "#12122a", "#2a2a5a", "#e0e0ff", "#7a7aaa")
    _paste_chart(draw, canvas, chart_path, (44, 236, 1036, 948), "#08081a", "#1a1a4a")

    headline_text = getattr(item, 'headline', f"อัปเดตหุ้น **{item.symbol}** วันนี้")
    draw_rich_text_wrapped(draw, headline_text, (74, 1000), headline_font, "#e0e0ff", "#FFC107", 932, line_gap=16)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_path, "PNG", optimize=True)
    return output_path


def compose_canvas_clean(item, chart_path, logo_bytes, output_path):
    """Style 3: Clean white — light background, professional minimal look."""
    from PIL import Image, ImageDraw
    canvas = Image.new("RGB", CANVAS_SIZE, "#f5f5f5")
    draw = ImageDraw.Draw(canvas)

    accent = BULL if item.pct_change >= 0 else BEAR
    draw.rectangle((0, 0, 1080, 8), fill=accent)

    title_font = load_font("Kanit-Bold.ttf", 66)
    company_font = load_font("Prompt-Regular.ttf", 30)
    metric_font = load_font("Kanit-Bold.ttf", 60)
    price_font = load_font("Prompt-Bold.ttf", 34)
    headline_font = load_font("Kanit-Bold.ttf", 56)

    draw.rounded_rectangle((44, 44, 1036, 216), radius=12, fill="#ffffff", outline="#e0e0e0", width=2)
    logo = image_from_logo_bytes(logo_bytes, item.symbol)
    canvas.paste(logo, (74, 70), logo)
    draw.text((218, 68), item.symbol, font=title_font, fill="#1a1a2e")
    company = item.company_name[:34] + ("..." if len(item.company_name) > 34 else "")
    draw.text((222, 145), company, font=company_font, fill="#666680")
    metric = f"+{item.pct_change:.2f}%" if item.pct_change >= 0 else f"{item.pct_change:.2f}%"
    metric_bbox = draw.textbbox((0, 0), metric, font=metric_font)
    draw.text((1002 - (metric_bbox[2] - metric_bbox[0]), 68), metric, font=metric_font, fill=accent)
    price = f"${item.price:,.2f}"
    price_bbox = draw.textbbox((0, 0), price, font=price_font)
    draw.text((1002 - (price_bbox[2] - price_bbox[0]), 145), price, font=price_font, fill="#1a1a2e")

    _paste_chart(draw, canvas, chart_path, (44, 236, 1036, 948), "#ffffff", "#e0e0e0")

    headline_text = getattr(item, 'headline', f"อัปเดตหุ้น **{item.symbol}** วันนี้")
    draw_rich_text_wrapped(draw, headline_text, (74, 1000), headline_font, "#1a1a2e", "#FFC107", 932, line_gap=16)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_path, "PNG", optimize=True)
    return output_path


def compose_canvas_bold(item, chart_path, logo_bytes, output_path):
    """Style 4: Bold headline — large text on top, smaller chart below. Impact-first design."""
    from PIL import Image, ImageDraw
    canvas = Image.new("RGB", CANVAS_SIZE, "#111111")
    draw = ImageDraw.Draw(canvas)

    accent = BULL if item.pct_change >= 0 else BEAR
    # Bold side accent bar
    draw.rectangle((0, 0, 12, 1350), fill=accent)
    draw.rectangle((1068, 0, 1080, 1350), fill=accent)

    title_font = load_font("Kanit-Bold.ttf", 72)
    company_font = load_font("Prompt-Regular.ttf", 28)
    metric_font = load_font("Kanit-Bold.ttf", 80)
    price_font = load_font("Prompt-Bold.ttf", 36)
    headline_font = load_font("Kanit-Bold.ttf", 64)

    # Large % change at top
    metric = f"+{item.pct_change:.2f}%" if item.pct_change >= 0 else f"{item.pct_change:.2f}%"
    metric_bbox = draw.textbbox((0, 0), metric, font=metric_font)
    metric_w = metric_bbox[2] - metric_bbox[0]
    draw.text(((1080 - metric_w) // 2, 50), metric, font=metric_font, fill=accent)

    # Symbol + company below
    sym_bbox = draw.textbbox((0, 0), item.symbol, font=title_font)
    sym_w = sym_bbox[2] - sym_bbox[0]
    draw.text(((1080 - sym_w) // 2, 155), item.symbol, font=title_font, fill="#ffffff")
    company = item.company_name[:40] + ("..." if len(item.company_name) > 40 else "")
    comp_bbox = draw.textbbox((0, 0), company, font=company_font)
    comp_w = comp_bbox[2] - comp_bbox[0]
    draw.text(((1080 - comp_w) // 2, 240), company, font=company_font, fill="#888888")

    # Price
    price = f"${item.price:,.2f}"
    price_bbox2 = draw.textbbox((0, 0), price, font=price_font)
    price_w = price_bbox2[2] - price_bbox2[0]
    draw.text(((1080 - price_w) // 2, 285), price, font=price_font, fill="#cccccc")

    # Chart (smaller, in middle)
    _paste_chart(draw, canvas, chart_path, (44, 350, 1036, 880), "#0a0a0a", "#333333")

    # Headline at bottom
    headline_text = getattr(item, 'headline', f"อัปเดตหุ้น **{item.symbol}** วันนี้")
    draw_rich_text_wrapped(draw, headline_text, (74, 930), headline_font, "#ffffff", "#FFC107", 932, line_gap=16)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(output_path, "PNG", optimize=True)
    return output_path


CANVAS_STYLES = {
    "viral": compose_canvas_viral,
    "classic": compose_canvas_classic,
    "neon": compose_canvas_neon,
    "clean": compose_canvas_clean,
    "bold": compose_canvas_bold,
}




def dropbox_client(creds: DropboxCreds):
    import dropbox

    if creds.refresh_token and creds.app_key and creds.app_secret:
        return dropbox.Dropbox(
            oauth2_refresh_token=creds.refresh_token,
            app_key=creds.app_key,
            app_secret=creds.app_secret,
        )
    if creds.access_token:
        return dropbox.Dropbox(creds.access_token)
    return None


def to_dropbox_raw_url(url: str) -> str:
    if not url:
        return ""
    if "?dl=0" in url:
        return url.replace("?dl=0", "?raw=1")
    if "&dl=0" in url:
        return url.replace("&dl=0", "&raw=1")
    base = url.split("?")[0]
    return f"{base}?raw=1"


def upload_to_dropbox(local_path: Path, dropbox_folder: str, creds: DropboxCreds) -> tuple[str, str]:
    import dropbox
    from dropbox.files import WriteMode
    from dropbox.sharing import RequestedVisibility, SharedLinkSettings

    dbx = dropbox_client(creds)
    if dbx is None:
        return "", ""

    clean_folder = ("/" + dropbox_folder.strip("/")).replace("//", "/")
    dropbox_path = f"{clean_folder}/{local_path.name}"
    with local_path.open("rb") as f:
        dbx.files_upload(f.read(), dropbox_path, mode=WriteMode("overwrite"), mute=True)

    try:
        link = dbx.sharing_create_shared_link_with_settings(
            dropbox_path, settings=SharedLinkSettings(requested_visibility=RequestedVisibility.public)
        ).url
    except dropbox.exceptions.ApiError:
        links = dbx.sharing_list_shared_links(path=dropbox_path, direct_only=True).links
        link = links[0].url if links else ""

    return dropbox_path, to_dropbox_raw_url(link)


def write_csv(items: list[StockItem], config_ref: str, csv_path: Path, run_date: str) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["Date", "Symbol", "Caption", "Image_URL", "Config_Ref"])
        writer.writeheader()
        for item in items:
            writer.writerow(
                {
                    "Date": run_date,
                    "Symbol": item.symbol,
                    "Caption": item.caption,
                    "Image_URL": item.dropbox_url,
                    "Config_Ref": config_ref,
                }
            )


def safe_file_part(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip())[:80] or "stock"


def generate_summary_post(items: list[StockItem], creds: AppCreds, model: str, mode: str, run_date: str) -> str:
    """Generate a summary/overview post that covers all stocks in the batch."""
    mode_label = {
        "gainers": "หุ้นพุ่งแรงวันนี้",
        "losers": "หุ้นร่วงหนักวันนี้",
        "low_pe": "หุ้น Value น่าจับตา",
        "trending": "หุ้นซื้อขายเดือดวันนี้",
    }.get(mode, "หุ้นน่าจับตาวันนี้")

    stock_list = "\n".join(
        f"- {item.symbol} ({item.company_name}): {item.pct_change:+.2f}% ราคา ${item.price:.2f} วอลุ่ม {item.volume:,}"
        for item in items
    )

    prompt = (
        f"You are a professional Thai stock analyst writing a Facebook page post.\n"
        f"Today is {run_date}. Topic: {mode_label}\n\n"
        f"Here are the stocks covered today:\n{stock_list}\n\n"
        f"Write a MAIN SUMMARY POST in Thai that will be the FIRST thing readers see.\n"
        f"This post introduces all {len(items)} stocks as a group.\n\n"
        f"Structure:\n"
        f"1. Eye-catching opening line with emoji (e.g. 🔥📊🚀)\n"
        f"2. Brief overview of today's market theme (2-3 sentences)\n"
        f"3. Quick bullet list of all stocks with 1-line summary each\n"
        f"4. A teaser line like 'ดูรายละเอียดวิเคราะห์แต่ละตัวในคอมเมนต์ด้านล่าง 👇'\n"
        f"5. Relevant hashtags (5-8 tags)\n\n"
        f"Keep it concise (150-250 words). Use Thai mixed with English terms.\n"
        f"Make it feel urgent, professional, and scroll-stopping.\n"
        f"Return ONLY the post text, nothing else."
    )

    if creds.openrouter_key:
        try:
            return chat_completion(
                api_key=creds.openrouter_key,
                base_url="https://openrouter.ai/api/v1/chat/completions",
                model=model,
                prompt=prompt,
                headers={"HTTP-Referer": "http://localhost", "X-Title": "BulkVideoCreatorApp"},
            )
        except Exception as exc:
            print(f"[WARN] Summary post AI failed: {exc}", file=sys.stderr)

    if creds.openai_key:
        try:
            return chat_completion(
                api_key=creds.openai_key,
                base_url="https://api.openai.com/v1/chat/completions",
                model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
                prompt=prompt,
                headers={},
            )
        except Exception as exc:
            print(f"[WARN] Summary post OpenAI failed: {exc}", file=sys.stderr)

    # Fallback: generate without AI
    lines = [f"📊 {mode_label} — {run_date}\n"]
    for item in items:
        arrow = "🟢" if item.pct_change >= 0 else "🔴"
        lines.append(f"{arrow} {item.symbol} ({item.company_name}) {item.pct_change:+.2f}% ราคา ${item.price:.2f}")
    lines.append(f"\nดูรายละเอียดวิเคราะห์แต่ละตัวในคอมเมนต์ด้านล่าง 👇")
    lines.append(f"\n#หุ้นวันนี้ #StockAnalysis #TopGainers #การลงทุน #หุ้นอเมริกา")
    return "\n".join(lines)


def run(args: argparse.Namespace) -> int:
    creds = load_credentials(args.profile)
    spreadsheet_path = Path(args.spreadsheet).expanduser().resolve()
    if not spreadsheet_path.exists():
        raise FileNotFoundError(f"Spreadsheet not found: {spreadsheet_path}")

    config_ref = read_p2(spreadsheet_path, args.sheet)
    config = parse_p2_config(config_ref)
    if not config_ref:
        print("[WARN] Cell P2 is empty. Continuing with Yahoo day gainers and blank Config_Ref.", file=sys.stderr)

    run_date = args.date or date.today().isoformat()
    output_dir = Path(args.output_dir).expanduser().resolve() / run_date
    chart_dir = output_dir / "charts"
    output_dir.mkdir(parents=True, exist_ok=True)
    chart_dir.mkdir(parents=True, exist_ok=True)

    print(f"[INFO] P2 Config_Ref: {config_ref or '(empty)'}")
    print(f"[INFO] Mode: {args.mode}")
    print(f"[INFO] Canvas style: {args.canvas_style}")
    print(f"[INFO] Image ratio: {args.image_ratio}")
    print(f"[INFO] Headline color set: {args.headline_theme}")
    print(f"[INFO] Meme Sticker overlay: {'ON' if args.meme_overlay else 'OFF'}")
    if args.meme_overlay:
        print(f"[INFO] Meme source: {args.meme_source}")
    if args.page_credit:
        print(f"[INFO] Page credit: {args.page_credit}")
    print(f"[INFO] Output folder: {output_dir}")

    items = get_stock_universe_by_mode(config, args.limit, args.scan_count, args.mode)
    if not items:
        raise RuntimeError("No stocks to process. Check P2 symbols/sector or the Yahoo screener response.")

    dropbox_folder = f"{args.dropbox_folder.rstrip('/')}/{run_date}"
    processed: list[StockItem] = []
    for index, item in enumerate(items, start=1):
        print(f"[INFO] ({index}/{len(items)}) Processing {item.symbol} {item.company_name}")
        setattr(item, "overlay_variant", index)
        item.news = fetch_news(item, creds)
        raw_caption = generate_caption(item, creds, args.model, args.mode)
        
        import re
        img_match = re.search(r'IMAGE_TEXT:\s*(.*?)\s*(?:POST_CAPTION:|$)', raw_caption, re.DOTALL)
        post_match = re.search(r'POST_CAPTION:\s*(.*)', raw_caption, re.DOTALL)
        
        if img_match:
            item.headline = img_match.group(1).strip()
        else:
            item.headline = f"อัปเดตหุ้น **{item.symbol}** วันนี้"
        item.headline = polish_image_headline(item, args.mode)
            
        if post_match:
            item.caption = post_match.group(1).strip()
        else:
            item.caption = raw_caption.strip()
        item.caption = append_news_sources(item.caption, item)
        news_card_copy = generate_news_overlay_copy(item, creds, args.model)
        if news_card_copy:
            setattr(item, "news_card_label", news_card_copy[0])
            setattr(item, "news_card_detail", news_card_copy[1])
            print(f"[INFO] News card copy for {item.symbol}: {news_card_copy[0]} / {news_card_copy[1]}")
        if args.meme_overlay:
            meme_title, meme_subtitle = generate_meme_copy(item, creds, args.model)
            setattr(item, "meme_title", meme_title)
            setattr(item, "meme_subtitle", meme_subtitle)
            print(f"[INFO] Meme copy for {item.symbol}: {meme_title} / {meme_subtitle}")
            
        chart_path = generate_chart(item, chart_dir / f"{safe_file_part(item.symbol)}_chart.png")
        logo_bytes = fetch_logo(item)
        image_path = output_dir / f"{run_date}_{safe_file_part(item.symbol)}.png"
        giphy_key = args.giphy_api_key or creds.giphy_api_key
        item.image_path = compose_canvas(
            item,
            chart_path,
            logo_bytes,
            image_path,
            style=args.canvas_style,
            image_ratio=args.image_ratio,
            headline_theme=args.headline_theme,
            meme_overlay=args.meme_overlay,
            page_credit=args.page_credit,
            meme_source=args.meme_source,
            giphy_api_key=giphy_key,
            local_meme_path=args.local_meme_path,
        )

        if not args.skip_dropbox:
            try:
                _, direct_url = upload_to_dropbox(image_path, dropbox_folder, creds.dropbox)
                item.dropbox_url = direct_url
                if direct_url:
                    print(f"[INFO] Dropbox raw URL: {direct_url}")
                elif args.require_dropbox:
                    raise RuntimeError("Dropbox credentials are missing.")
                else:
                    print("[WARN] Dropbox credentials missing; Image_URL will be blank.", file=sys.stderr)
            except Exception as exc:
                if args.require_dropbox:
                    raise
                print(f"[WARN] Dropbox upload failed for {item.symbol}: {exc}", file=sys.stderr)

        processed.append(item)
        if args.pause_seconds > 0 and index < len(items):
            time.sleep(args.pause_seconds)

    # Generate summary post for the entire batch
    summary_post = generate_summary_post(processed, creds, args.model, args.mode, run_date)
    summary_path = output_dir / "summary_post.txt"
    summary_path.write_text(summary_post, encoding="utf-8")
    print(f"[INFO] Summary post saved: {summary_path}")

    csv_path = Path(args.csv_path).expanduser().resolve() if args.csv_path else output_dir / "content_factory_post.csv"
    write_csv(processed, config_ref, csv_path, run_date)
    print(f"[DONE] Wrote {csv_path}")
    return 0


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate top-gainers social images, upload to Dropbox, and write CSV.")
    parser.add_argument("--spreadsheet", required=True, help="Master .xlsx/.xlsm/.csv config file. Cell P2 is always used.")
    parser.add_argument("--sheet", default="", help="Worksheet name for Excel files. Defaults to the active sheet.")
    parser.add_argument("--limit", type=int, default=5, help="Number of stocks/images to generate.")
    parser.add_argument("--scan-count", type=int, default=150, help="Yahoo day-gainer candidates to scan for sector filters.")
    parser.add_argument("--date", default="", help="Generation date, YYYY-MM-DD. Defaults to today.")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_ROOT), help="Local output root. Date folder is appended.")
    parser.add_argument("--dropbox-folder", default=DEFAULT_DROPBOX_ROOT, help="Dropbox root folder. Date folder is appended.")
    parser.add_argument("--csv-path", default="", help="Optional explicit CSV output path.")
    parser.add_argument("--profile", default="", help="Optional api_profiles.json profile id/name.")
    parser.add_argument("--model", default=os.getenv("OPENROUTER_MODEL", "google/gemini-2.5-flash"), help="LLM model.")
    parser.add_argument("--skip-dropbox", action="store_true", help="Generate local files and CSV without uploading.")
    parser.add_argument("--require-dropbox", action="store_true", help="Fail if Dropbox upload cannot complete.")
    parser.add_argument("--pause-seconds", type=float, default=0.5, help="Delay between stocks to be gentle with APIs.")
    parser.add_argument("--mode", default="gainers",
                        choices=["gainers", "losers", "low_pe", "trending"],
                        help="Screening mode: gainers=top gainers, losers=top losers, low_pe=low P/E value stocks, trending=most active/news-heavy")
    parser.add_argument("--canvas-style", default="viral",
                        choices=["viral", "classic", "neon", "clean", "bold"],
                        help="Visual style for the generated canvas images.")
    parser.add_argument("--image-ratio", default="default",
                        choices=["default", "square"],
                        help="Image size ratio: default=existing 1080x1350 layout, square=1:1 1080x1080 viral layout.")
    parser.add_argument("--headline-theme", default="classic",
                        choices=sorted(HEADLINE_COLOR_THEMES.keys()),
                        help="Headline highlight color set for the viral canvas.")
    parser.add_argument("--meme-overlay", action="store_true", help="Add a small meme-style reaction sticker to the viral canvas.")
    parser.add_argument("--page-credit", default="", help="Optional page name/credit text rendered at the bottom of the viral canvas.")
    parser.add_argument("--meme-source", default="local", choices=["local", "giphy"], help="Meme source: local=file/local sticker, giphy=GIPHY API still image.")
    parser.add_argument("--giphy-api-key", default="", help="Optional GIPHY API key override. Falls back to active API profile or GIPHY_API_KEY.")
    parser.add_argument("--local-meme-path", default="", help="Optional local meme image path for meme-source=local.")
    return parser


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()
    try:
        return run(args)
    except KeyboardInterrupt:
        print("\n[STOPPED] Interrupted by user.", file=sys.stderr)
        return 130
    except Exception as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
